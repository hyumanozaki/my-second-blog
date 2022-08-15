from django.shortcuts import render
from django.utils import timezone
from .models import Post
from .models import MBO22,ADC22,RHDT,PDI22#, PDI22G
#from .models import MBO
from .forms import PostForm
from .forms import MBO22Q1Form,MBO22Q2Form,MBO22Q3Form,MBO22Q4Form,ADC22CForm,ADC22AOForm,PDI22Form,DeptForm,uploadForm, readForm
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.views.generic import TemplateView #テンプレートタグ
#from .forms import AccountForm#, AddAccountForm #ユーザーアカウントフォー
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate, login, logout
from django.core.mail import send_mail
from datetime import datetime
from django.views.generic import CreateView
from django.urls import reverse_lazy
import sys
import os
import xlsxwriter
import io
import openpyxl as op
from apscheduler.schedulers.background import BackgroundScheduler
#import numpy as np
#import pandas as pd
#from user.models import User, Order

User = get_user_model()

def Login(request):
    # POST
    if request.method == 'POST':
        # フォーム入力のユーザーID・パスワード取得
        ID = request.POST.get('userid')
        Pass = request.POST.get('password')

        # Djangoの認証機能
        user = authenticate(username=ID, password=Pass)

        # ユーザー認証
        if user:
            #ユーザーアクティベート判定
            if user.is_active:
                # ログイン
                login(request,user)
                # ホームページ遷移
                return HttpResponseRedirect(reverse('home'))
                #return render(request, "blog/home.html",user)
            else:
                # アカウント利用不可
                return HttpResponse("アカウントが有効ではありません")
        # ユーザー認証失敗
        else:
            return HttpResponse("endereço de email ou senha está errado")
    # GET
    else:
        return render(request, 'blog/login.html')
    
#    numero = Account.objects.all()
#    for item in numero:
#        MBO = AccountForm.objects.userid(item)
#        MBO.save()


def homeA(request):

    AA=''
    BB=[]
    CC=''
    DD=[]

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    result = ''
    result2 = ''
    result3 = ''
    result4 = ''
    result5 = ''
    result6 = ''
    result7 = ''
    result8 = ''
    object_list = User.objects.all()
    for i in object_list:
        result = i.email
        result2 += result
        result2 += ' '
        result3 = i.username
        result4 += result3
        result4 += ' '
        result6 = str(i.id)
        result7 += result6
        result7 += ' '
    result2=result2.split()#email de todos
    result4=result4.split()#name de todos
    result7=result7.split()#id de tods
    k=int(len(result2))
    for j in range(k):
        if (str(result2[j])==str(request.user)):
            result5+=str(result4[j])
            result5+=' '
            result8+=str(result7[j])
            result8+=' '
            ID=int(result7[j])

            AA+='#'
            AA+='^'
            AA+=str(result4[j])
            AA+='^'
            AA+='#'
            AA+='^'

            try:             

                DD=RHDT.objects.values_list('MBOTIME22').get(user=resultb6)
                DD1=MBO22.objects.values_list('MBO22Q1D').get(user=ID)
                DD2=MBO22.objects.values_list('MBO22Q2D').get(user=ID)
                DD3=MBO22.objects.values_list('MBO22Q3D').get(user=ID)
                DD4=MBO22.objects.values_list('MBO22Q4D').get(user=ID)
                if int(DD[0])==1:
                    session='(define meta e peso% de 2022)'
                    if int(DD1[0])==0:
                        CC+=str(result4[j])
                        CC+='^'
                elif int(DD[0])==2:
                    session='(primeiro alinhamento de 2022)'
                    if int(DD2[0])==0:
                        CC+=str(result4[j])
                        CC+='^'
                elif int(DD[0])==3:
                    session='(segundo alinhamento de 2022)'
                    if int(DD3[0])==0:
                        CC+=str(result4[j])
                        CC+='^'
                elif int(DD[0])==4:
                    session='(avaliação final de fim do ano e pontuação de 2022)'
                    if int(DD4[0])==0:
                        CC+=str(result4[j])
                        CC+='^'

                BB=MBO22.objects.values_list('MBO22A1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22AP').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22A2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22A3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22A4').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22AR').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=MBO22.objects.values_list('MBO22B1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22BP').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22B2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22B3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22B4').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22BR').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=MBO22.objects.values_list('MBO22C1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22CP').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22C2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22C3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22C4').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22CR').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=MBO22.objects.values_list('MBO22D1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22DP').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22D2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22D3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22D4').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22DR').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=MBO22.objects.values_list('MBO22E1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22EP').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22E2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22E3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22E4').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22ER').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=MBO22.objects.values_list('MBO22F1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22FP').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22F2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22F3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22F4').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22FR').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=MBO22.objects.values_list('MBO22G1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22GP').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22G2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22G3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22G4').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=MBO22.objects.values_list('MBO22GR').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
            except:
                AA+='não tem MBO'
                AA+='^'
    result5=result5.split()
    result8=result8.split()
    AA=AA.split('^')
    CC=CC.split('^')
    del AA[-1]
    params = {
        "UserID":request.user,
        "time":AA,
        "C1":result5,
        "A":result4,
        "C3":result4,
        "UserID":request.user,
        "CC":CC,
        "session":session,
        }
    return render(request, "blog/homeA.html",params)

def homeB(request):

    AA=''
    BB=[]
    CC=''
    DD=[]

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])


    result = ''
    result2 = ''
    result3 = ''
    result4 = ''
    result5 = ''
    result6 = ''
    result7 = ''
    result8 = ''
    object_list = User.objects.all()
    for i in object_list:
        result = i.email
        result2 += result
        result2 += ' '
        result3 = i.username
        result4 += result3
        result4 += ' '
        result6 = str(i.id)
        result7 += result6
        result7 += ' '
    result2=result2.split()
    result4=result4.split()
    result7=result7.split()
    k=int(len(result2))
    for j in range(k):
        if (str(result2[j])==str(request.user)):
            result5+=str(result4[j])
            result5+=' '
            result8+=str(result7[j])
            result8+=' '
            ID = int(result7[j])

            AA+='#'
            AA+='^'
            AA+=str(result4[j])
            AA+='^'
            AA+='#'
            AA+='^'
            try:
                DD=ADC22.objects.values_list('ADC22D').get(user=ID)
                if int(DD[0]) == 0:
                    CC+=str(result4[j])
                    CC+='^'
                BB=RHDT.objects.values_list('ADC22G1').get(user=resultb6)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G1C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G1A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G1OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G1O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=RHDT.objects.values_list('ADC22G2').get(user=resultb6)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G2C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G2A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G2OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G2O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=RHDT.objects.values_list('ADC22G3').get(user=resultb6)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G3C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G3A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G3OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G3O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=RHDT.objects.values_list('ADC22G4').get(user=resultb6)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G4C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G4A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G4OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G4O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=RHDT.objects.values_list('ADC22G5').get(user=resultb6)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G5C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G5A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G5OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G5O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=RHDT.objects.values_list('ADC22G6').get(user=resultb6)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G6C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G6A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G6OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G6O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=RHDT.objects.values_list('ADC22G7').get(user=resultb6)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G7C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G7A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G7OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22G7O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E1C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E1A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E1OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E1O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E2C').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E2A').get(user=ID)
                AA+=str(BB[0])
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E2OC').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=ADC22.objects.values_list('ADC22E2O').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
            except:
                AA+='não tem avaliação de competencia'
                AA+='^'
    AA=AA.split('^')
    CC=CC.split('^')
    result5=result5.split()
    result8=result8.split()
    del AA[-1]
    params = {
        "UserID":request.user,
        "time":AA,
        "C1":result5,
        "C2":result8,
        "C3":result4,
        "UserID":request.user,
        "CC":CC,
        }
    return render(request, "blog/homeB.html",params)

def homeC(request):

    AA=''
    BB=[]
    CC=''
    DD=[]

    result = ''
    result2 = ''
    result3 = ''
    result4 = ''
    result5 = ''
    result6 = ''
    result7 = ''
    result8 = ''
    object_list = User.objects.all()
    for i in object_list:
        result = i.email
        result2 += result
        result2 += ' '
        result3 = i.username
        result4 += result3
        result4 += ' '
        result6 = str(i.id)
        result7 += result6
        result7 += ' '
    result2=result2.split()
    result4=result4.split()
    result7=result7.split()
    k=int(len(result2))
    for j in range(k):
        if (str(result2[j])==str(request.user)):
            result5+=str(result4[j])
            result5+=' '
            result8+=str(result7[j])
            result8+=' '
            ID=int(result7[j])

            AA+='#'
            AA+='^'
            AA+=str(result4[j])
            AA+='^'
            AA+='#'
            AA+='^'
            try:
                DD=PDI22.objects.values_list('PDI22D').get(user=ID)
                if int(DD[0]) == 0:
                    CC+=str(result4[j])
                    CC+='^'
                BB=PDI22.objects.values_list('PDI22G1C').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G1PD').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G2C').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G2PD').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G3C').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G3PD').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22G3').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=PDI22.objects.values_list('PDI22E1C').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22E1PD').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22E1').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
                BB=PDI22.objects.values_list('PDI22E2C').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22E2PD').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                BB=PDI22.objects.values_list('PDI22E2').get(user=ID)
                AA+=BB[0]
                BB=[]
                AA+='^'
                AA+='#'
                AA+='^'
            except:
                AA+='não tem PDI'
                AA+='^'
    AA=AA.split('^')
    CC=CC.split('^')
    result5=result5.split()
    result8=result8.split()
    del AA[-1]
    params = {
        "UserID":request.user,
        "time":AA,
        "C1":result5,
        "C2":result8,
        "C3":result4,
        "UserID":request.user,
        "CC":CC,
        }
    return render(request, "blog/homeC.html",params)


@login_required
def PDI2(request, name):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])


    colaborador = name#request.GET['name']

    result = ''
    result2 = ''
    result3 = ''
    result4 = ''
    result5 = ''
    result6 = 0
    result7 = ''
    result8 = ''
    result9 = ''
    result10 = ''
    result11 = ''
    result12 = ''
    result13 = ''
    result14 = ''
    result15 = ''
    object_list = User.objects.all()
    for i in object_list:
        result = i.username
        result2 += result
        result2 += ' '
        result3 = str(i.id)
        result4 += result3
        result4 += ' '
        result7 = i.email
        result8 += result7
        result8 += ' '
        result10 = i.first_name
        result11 += result10
        result11 += '^'
        result12 = i.last_name
        result13 += result12
        result13 += '^'
    result2=result2.split()
    result4=result4.split()
    result8=result8.split()
    result11=result11.split('^')
    result13=result13.split('^')
    k=int(len(result2))
    for j in range(k):
        if (str(result2[j])==colaborador):
            result6=int(result4[j])
            result9=str(result8[j])
            result14=str(result11[j])
            result15=str(result13[j])
    if result9 == str(request.user) :
        reject = 0
    else :
        reject = 1
#    params = {"UserID":request.user,"data1":MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=colaborador),"data2":MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=colaborador),"data3":MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=colaborador),"data4":MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=colaborador),"data5":MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4').get(user=colaborador),"data6":MBO22.objects.values_list('MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=colaborador)}

    porta=PDI22.objects.values_list('PDI22C').get(user=result6)
    porta2=PDI22.objects.values_list('PDI22A').get(user=result6) 

    if (int(porta[0])==0):
        time=0
    if (int(porta[0])==1):
        if (int(porta2[0])==0):
            time=1
        else :
            time=2

    today = datetime.now()
    Y = today.year
    M = today.month
    D = today.day



    Glist=[]
    Glist.append('resultado')
    Glist+=ADC22.objects.values_list('ADC22G1C').get(user=result6),
    Glist+=ADC22.objects.values_list('ADC22G1A').get(user=result6),
    Glist.append('cliente')
    Glist+=ADC22.objects.values_list('ADC22G2C').get(user=result6),
    Glist+=ADC22.objects.values_list('ADC22G2A').get(user=result6),
    Glist.append('busca')
    Glist+=ADC22.objects.values_list('ADC22G3C').get(user=result6),
    Glist+=ADC22.objects.values_list('ADC22G3A').get(user=result6),
    Glist.append('liderança')
    Glist+=ADC22.objects.values_list('ADC22G4C').get(user=result6),
    Glist+=ADC22.objects.values_list('ADC22G4A').get(user=result6),
    Glist.append('senso')
    Glist+=ADC22.objects.values_list('ADC22G5C').get(user=result6),
    Glist+=ADC22.objects.values_list('ADC22G5A').get(user=result6),
    Glist.append('comunicação')
    Glist+=ADC22.objects.values_list('ADC22G6C').get(user=result6),
    Glist+=ADC22.objects.values_list('ADC22G6A').get(user=result6),
    Glist.append('relacionamento')
    Glist+=ADC22.objects.values_list('ADC22G7C').get(user=result6),
    Glist+=ADC22.objects.values_list('ADC22G7A').get(user=result6),

    G1C=[]
    G2C=[]
    G3C=[]
    G1A=[]
    G2A=[]
    G3A=[]

    E1C=[]
    E2C=[]
    E1A=[]
    E2A=[]

    for i in range(7):
        j=i*3
        k=j+1
        l=k+1

        if str(Glist[j]) in str(PDI22.objects.values_list('PDI22G1C').get(user=result6)).lower():
            G1C+=Glist[k]
            G1A+=Glist[l]
        if str(Glist[j]) in str(PDI22.objects.values_list('PDI22G2C').get(user=result6)).lower():
            G2C+=Glist[k]
            G2A+=Glist[l]
        if str(Glist[j]) in str(PDI22.objects.values_list('PDI22G3C').get(user=result6)).lower():
            G3C+=Glist[k]
            G3A+=Glist[l]

    E1C+=ADC22.objects.values_list('ADC22E1C').get(user=result6)
    E1A+=ADC22.objects.values_list('ADC22E1A').get(user=result6)
    E2C+=ADC22.objects.values_list('ADC22E2C').get(user=result6)
    E2A+=ADC22.objects.values_list('ADC22E2A').get(user=result6)

    if (G1C==[]):
        data11=''
    else :
        data11=str(G1C[0])
    if (G2C==[]):
        data21=''
    else :
        data21=G2C[0]
    if (G3C==[]):
        data31=''
    else :
        data31=G3C[0]
    if (G1A==[]):
        data12=''
    else :
        data12=G1A[0]
    if (G2A==[]):
        data22=''
    else :
        data22=G2A[0]
    if (G3A==[]):
        data32=''
    else :
        data32=G3A[0]


    task = get_object_or_404(PDI22, user=result6)
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = xlsxwriter.Workbook(output)
            ws = book.add_worksheet('PDI dele(a)')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'foco de competencia geral', format)
            ws.write(1,2,'auto avaliação', format)
            ws.write(1,3,'avaliação por avaliador', format)
            ws.write(1,4,'ponto a desenvolver', format)
            ws.write(1,5,'plano de desenvolvimento individual', format)
            ws.write(2,1,task.PDI22G1C, format)
            ws.write(2,2,data11, format)
            ws.write(2,3,data12, format)
            ws.write(2,4,task.PDI22G1PD, format)
            ws.write(2,5,task.PDI22G1, format)
            ws.write(3,1,task.PDI22G2C, format)
            ws.write(3,2,data21, format)
            ws.write(3,3,data22, format)
            ws.write(3,4,task.PDI22G2PD, format)
            ws.write(3,5,task.PDI22G2, format)
            ws.write(4,1,task.PDI22G3C, format)
            ws.write(4,2,data31, format)
            ws.write(4,3,data32, format)
            ws.write(4,4,task.PDI22G3PD, format)
            ws.write(4,5,task.PDI22G3, format)
            ws.write(6,1,'competencia especifica', format)
            ws.write(6,2,'auto avaliação', format)
            ws.write(6,3,'avaliação por avaliador', format)
            ws.write(6,4,'ponto a desenvolver', format)
            ws.write(6,5,'plano de desenvolvimento individual', format)
            ws.write(7,1,result14, format)
            ws.write(7,2,str(E1C[0]), format)
            ws.write(7,3,str(E1A[0]), format)
            ws.write(7,4,task.PDI22E1PD, format)
            ws.write(7,5,task.PDI22E1, format)
            ws.write(8,1,result15, format)
            ws.write(8,2,str(E2C[0]), format)
            ws.write(8,3,str(E2A[0]), format)
            ws.write(8,4,task.PDI22E2PD, format)
            ws.write(8,5,task.PDI22E2, format)
            ws.set_column('B:B',20)
            ws.set_column('C:D',11)
            ws.set_column('E:F',40)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            book.close()
            output.seek(0)
            filename = 'PDI dele(a).xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response

        elif "send" in request.POST:
            new_status = 1
            task.PDI22A = new_status
            task.PDI22Y = Y
            task.PDI22M = M
            task.PDI22D = D
            task.save()
            time=2
            subject = "seu PDI foi aprovado pelo avaliador/a, parabens!"
            message = "seu PDI foi aprovado pelo avaliador/a, parabens!"
            from_email = 'sistema.rh@cosmotec.com.br'  # 送信
            #from_email = 'hyuma2331@gmail.com'  # 送信
            recipient_list = [colaborador]  # 宛先リスト
            send_mail(subject, message, from_email, recipient_list)
        elif "deny" in request.POST:
            new_status = 0
            task.PDI22C = new_status
            task.save()
            time=0
            subject = "seu avaliador/a pediu te para corrigir seu PDI"
            message = "seu avaliador/a pediu te para corrigir seu PDI, talvez tenha algo errado e tenha que corrigir, entre o sistema e corrija"
            from_email = 'sistema.rh@cosmotec.com.br'  # 送信
#            from_email = 'hyuma2331@gmail.com'  # 送信
            recipient_list = [colaborador]  # 宛先リスト
            send_mail(subject, message, from_email, recipient_list)

    params = {"UserID":request.user,
        "avaliado":colaborador,
        "data1":PDI22.objects.values_list('PDI22G1C','PDI22G1PD','PDI22G1').get(user=result6),
        "data11":data11,
        "data12":data12,
        "data2":PDI22.objects.values_list('PDI22G2C','PDI22G2PD','PDI22G2').get(user=result6),
        "data21":data21,
        "data22":data22,
        "data3":PDI22.objects.values_list('PDI22G3C','PDI22G3PD','PDI22G3').get(user=result6),
        "data31":data31,
        "data32":data32,
#        "data40":ADC22.objects.values_list('ADC22E1').get(user=result6),
        "data40":result14,
        "data4":PDI22.objects.values_list('PDI22E1PD','PDI22E1').get(user=result6),
        "data41":E1C[0],
        "data42":E1A[0],
#        "data50":ADC22.objects.values_list('ADC22E2').get(user=result6),
        "data50":result15,
        "data5":PDI22.objects.values_list('PDI22E2PD','PDI22E2').get(user=result6),
        "data51":E2C[0],
        "data52":E2A[0],
        "time":time,
        "reject":reject,
        }
    return render(request, "blog/PDI2.html", context=params)



@login_required
def Logout2(request, name):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    MBO22TIME=[]
    MBO22TIME+=RHDT.objects.values_list('MBOTIME22').get(user=resultb6)
    MBO22TIME=int(MBO22TIME[0])

    colaborador = name#request.GET['name']

    result = ''
    result2 = ''
    result3 = ''
    result4 = ''
    result5 = ''
    result6 = 0
    result7 = ''
    result8 = ''
    result9 = ''
    object_list = User.objects.all()
    for i in object_list:
        result = i.username
        result2 += result
        result2 += ' '
        result3 = str(i.id)
        result4 += result3
        result4 += ' '
        result7 = i.email
        result8 += result7
        result8 += ' '
    result2=result2.split()
    result4=result4.split()
    result8=result8.split()
    k=int(len(result2))
    for j in range(k):
        if (str(result2[j])==colaborador):
            result6=int(result4[j])
            result9=str(result8[j])
    if result9 == str(request.user) :
        reject = 0
    else :
        reject = 1
#    params = {"UserID":request.user,"data1":MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=colaborador),"data2":MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=colaborador),"data3":MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=colaborador),"data4":MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=colaborador),"data5":MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4').get(user=colaborador),"data6":MBO22.objects.values_list('MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=colaborador)}


    today = datetime.now()
    Y = today.year
    M = today.month
    D = today.day

    task = get_object_or_404(MBO22, user=result6)
    if (request.method == 'POST'):
        if "send" in request.POST:        
            new_status = 1
            if (MBO22TIME==1):
                task.MBO22Q1A = new_status
                task.MBO22Q1Y = Y
                task.MBO22Q1M = M
                task.MBO22Q1D = D
                task.save()
            if (MBO22TIME==2):
                task.MBO22Q2A = new_status
                task.MBO22Q2Y = Y
                task.MBO22Q2M = M
                task.MBO22Q2D = D
                task.save()
            if (MBO22TIME==3):
                task.MBO22Q3A = new_status
                task.MBO22Q3Y = Y
                task.MBO22Q3M = M
                task.MBO22Q3D = D
                task.save()
            if (MBO22TIME==4):
                task.MBO22Q4A = new_status
                task.MBO22Q4Y = Y
                task.MBO22Q4M = M
                task.MBO22Q4D = D
                task.save()
            subject = "seu MBO foi aprovado pelo avaliador/a, parabens!"
            message = "seu MBO foi aprovado pelo avaliador/a, parabens!"
            from_email = 'sistema.rh@cosmotec.com.br'  # 送信            
            recipient_list = [colaborador]  # 宛先リスト
            send_mail(subject, message, from_email, recipient_list)
        elif "deny" in request.POST:
            new_status = 0
            if (MBO22TIME==1):
                task.MBO22Q1 = new_status
                task.save()
            if (MBO22TIME==2):
                task.MBO22Q2 = new_status
                task.save()
            if (MBO22TIME==3):
                task.MBO22Q3 = new_status
                task.save()
            if (MBO22TIME==4):
                task.MBO22Q4 = new_status
                task.save()
            subject = "seu avaliador/a pediu te para corrigir seu MBO"
            message = "seu avaliador/a pediu te para corrigir seu MBO, talvez tenha algo errado e tenha que corrigir, entre o sistema e corrija"
#            from_email = 'hyuma2331@gmail.com'  # 送信
            from_email = 'sistema.rh@cosmotec.com.br'  # 送信
            recipient_list = [colaborador]  # 宛先リスト
            send_mail(subject, message, from_email, recipient_list)
        elif "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = xlsxwriter.Workbook(output)
            ws = book.add_worksheet('MBO dele(a)')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'descrição de meta', format)
            ws.write(1,2,'peso%', format)
            ws.write(1,3,'primeiro alinhamento', format)
            ws.write(1,4,'segundo alinhamento', format)
            ws.write(1,5,'avalaição de fim do ano', format)
            ws.write(1,6,'pontuação', format)
            ws.write(2,1,task.MBO22A1, format)
            ws.write(2,2,task.MBO22AP, format)
            ws.write(2,3,task.MBO22A2, format)
            ws.write(2,4,task.MBO22A3, format)
            ws.write(2,5,task.MBO22A4, format)
            ws.write(2,6,task.MBO22AR, format)
            ws.write(3,1,task.MBO22B1, format)
            ws.write(3,2,task.MBO22BP, format)
            ws.write(3,3,task.MBO22B2, format)
            ws.write(3,4,task.MBO22B3, format)
            ws.write(3,5,task.MBO22B4, format)
            ws.write(3,6,task.MBO22BR, format)
            ws.write(4,1,task.MBO22C1, format)
            ws.write(4,2,task.MBO22CP, format)
            ws.write(4,3,task.MBO22C2, format)
            ws.write(4,4,task.MBO22C3, format)
            ws.write(4,5,task.MBO22C4, format)
            ws.write(4,6,task.MBO22CR, format)
            ws.write(5,1,task.MBO22D1, format)
            ws.write(5,2,task.MBO22DP, format)
            ws.write(5,3,task.MBO22D2, format)
            ws.write(5,4,task.MBO22D3, format)
            ws.write(5,5,task.MBO22D4, format)
            ws.write(5,6,task.MBO22DR, format)
            ws.write(6,1,task.MBO22E1, format)
            ws.write(6,2,task.MBO22EP, format)
            ws.write(6,3,task.MBO22E2, format)
            ws.write(6,4,task.MBO22E3, format)
            ws.write(6,5,task.MBO22E4, format)
            ws.write(6,6,task.MBO22ER, format)
            ws.write(7,1,task.MBO22F1, format)
            ws.write(7,2,task.MBO22FP, format)
            ws.write(7,3,task.MBO22F2, format)
            ws.write(7,4,task.MBO22F3, format)
            ws.write(7,5,task.MBO22F4, format)
            ws.write(7,6,task.MBO22FR, format)
            ws.write(8,1,task.MBO22G1, format)
            ws.write(8,2,task.MBO22GP, format)
            ws.write(8,3,task.MBO22G2, format)
            ws.write(8,4,task.MBO22G3, format)
            ws.write(8,5,task.MBO22G4, format)
            ws.write(8,6,task.MBO22GR, format)
            ws.set_column('B:B',33)
            ws.set_column('C:C',11)
            ws.set_column('D:E',33)
            ws.set_column('F:F',11)
            ws.set_column('G:G',33)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(5,50)
            ws.set_row(6,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            book.close()
            output.seek(0)
            filename = 'MBO dele(a).xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response

    disp=0
    porta=MBO22.objects.values_list('MBO22Q1').get(user=result6)+MBO22.objects.values_list('MBO22Q2').get(user=result6)+MBO22.objects.values_list('MBO22Q3').get(user=result6)+MBO22.objects.values_list('MBO22Q4').get(user=result6)
    porta2=MBO22.objects.values_list('MBO22Q1A').get(user=result6)+MBO22.objects.values_list('MBO22Q2A').get(user=result6)+MBO22.objects.values_list('MBO22Q3A').get(user=result6)+MBO22.objects.values_list('MBO22Q4A').get(user=result6)
    if (MBO22TIME==1):
        if (int(porta[0])==1):
            if (int(porta2[0])==0):
                    time=1
            else:
                time=10
        else :
            time=0
    if (MBO22TIME==2):
        if (int(porta[1])==1):
            if (int(porta2[1])==0):
                time=2
            else:
                time=10
        else :
            time=0
    if (MBO22TIME==3):
        if (int(porta[2])==1):
            if (int(porta2[2])==0):
                time=3
            else:
                time=10
        else :
            time=0
    if (MBO22TIME==4):
        disp=1
        if (int(porta[3])==1):
            if (int(porta2[3])==0):
                time=4
            else:
                time=10
        else :
            time=0

    obj = MBO22.objects.get(user=result6)
    params = {
        "UserID":request.user,
        "avaliado":colaborador,
        "data1":MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=result6),
        "data2":MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=result6),
        "data3":MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=result6),
        "data4":MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=result6),
        "data5":MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4').get(user=result6),
        "data6":MBO22.objects.values_list('MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=result6),
        "time":time,"disp":disp,"reject" : reject,"result9" : result9,
        }      
    return render(request, "blog/Logout2.html", context=params)



#ログアウト
@login_required
def Logout(request):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    MBO22TIME=[]
    MBO22TIME+=RHDT.objects.values_list('MBOTIME22').get(user=resultb6)
    MBO22TIME=int(MBO22TIME[0])

    disp=0
    porta=[]
    porta+=MBO22.objects.values_list('MBO22Q1').get(user=request.user)
    porta+=MBO22.objects.values_list('MBO22Q2').get(user=request.user)
    porta+=MBO22.objects.values_list('MBO22Q3').get(user=request.user)
    porta+=MBO22.objects.values_list('MBO22Q4').get(user=request.user)
    if (MBO22TIME==1):
        if (int(porta[0])==0):
            time=1
        else :
            time=0
    if (MBO22TIME==2):
        if (int(porta[1])==0):
            time=2
        else :
            time=0
    if (MBO22TIME==3):
        if (int(porta[2])==0):
            time=3
        else :
            time=0
    if (MBO22TIME==4):
        disp=1
        if (int(porta[3])==0):
            time=4
        else :
            time=0

    task = get_object_or_404(MBO22, user=request.user)
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = xlsxwriter.Workbook(output)
            ws = book.add_worksheet('seu MBO')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'descrição de metas', format)
            ws.write(1,2,'peso%', format)
            ws.write(1,3,'primeiro alinhamento', format)
            ws.write(1,4,'segundo alinhamento', format)
            ws.write(1,5,'avaliação de fim do ano', format)
            ws.write(1,6,'pontuação', format)
            ws.write(2,1,task.MBO22A1, format)
            ws.write(2,2,task.MBO22AP, format)
            ws.write(2,3,task.MBO22A2, format)
            ws.write(2,4,task.MBO22A3, format)
            ws.write(2,5,task.MBO22A4, format)
            ws.write(2,6,task.MBO22AR, format)
            ws.write(3,1,task.MBO22B1, format)
            ws.write(3,2,task.MBO22BP, format)
            ws.write(3,3,task.MBO22B2, format)
            ws.write(3,4,task.MBO22B3, format)
            ws.write(4,1,task.MBO22C1, format)
            ws.write(4,2,task.MBO22CP, format)
            ws.write(4,3,task.MBO22C2, format)
            ws.write(4,4,task.MBO22C3, format)
            ws.write(5,1,task.MBO22D1, format)
            ws.write(5,2,task.MBO22DP, format)
            ws.write(5,3,task.MBO22D2, format)
            ws.write(5,4,task.MBO22D3, format)
            ws.write(6,1,task.MBO22E1, format)
            ws.write(6,2,task.MBO22EP, format)
            ws.write(6,3,task.MBO22E2, format)
            ws.write(6,4,task.MBO22E3, format)
            ws.write(7,1,task.MBO22F1, format)
            ws.write(7,2,task.MBO22FP, format)
            ws.write(7,3,task.MBO22F2, format)
            ws.write(7,4,task.MBO22F3, format)
            ws.write(8,1,task.MBO22G1, format)
            ws.write(8,2,task.MBO22GP, format)
            ws.write(8,3,task.MBO22G2, format)
            ws.write(8,4,task.MBO22G3, format)
            ws.write(3,5,task.MBO22B4, format)
            ws.write(3,6,task.MBO22BR, format)
            ws.write(4,5,task.MBO22C4, format)
            ws.write(4,6,task.MBO22CR, format)
            ws.write(5,5,task.MBO22D4, format)
            ws.write(5,6,task.MBO22DR, format)
            ws.write(6,5,task.MBO22E4, format)
            ws.write(6,6,task.MBO22ER, format)
            ws.write(7,5,task.MBO22F4, format)
            ws.write(7,6,task.MBO22FR, format)
            ws.write(8,5,task.MBO22G4, format)
            ws.write(8,6,task.MBO22GR, format)
            ws.set_column('B:B',33)
            ws.set_column('C:C',11)
            ws.set_column('D:F',33)
            ws.set_column('G:G',11)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(5,50)
            ws.set_row(6,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            book.close()
            output.seek(0)
            filename = 'seu MBO.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response
        elif "send" in request.POST:
            if (time==1):
                ptotal = MBO22.objects.values_list('MBO22AP').get(user=request.user)+MBO22.objects.values_list('MBO22BP').get(user=request.user)+MBO22.objects.values_list('MBO22CP').get(user=request.user)+MBO22.objects.values_list('MBO22DP').get(user=request.user)+MBO22.objects.values_list('MBO22EP').get(user=request.user)+MBO22.objects.values_list('MBO22FP').get(user=request.user)+MBO22.objects.values_list('MBO22GP').get(user=request.user)
                PT = int(ptotal[0])+int(ptotal[1])+int(ptotal[2])+int(ptotal[3])+int(ptotal[4])+int(ptotal[5])+int(ptotal[6])
                if (PT==100):
                    new_status = 1
                    task.MBO22Q1 = new_status
                    task.save()
                else :
                    time = 5
                    params = {"UserID":request.user,"data1":MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=request.user),"data2":MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=request.user),"data3":MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=request.user),"data4":MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=request.user),"data5":MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4').get(user=request.user),"data6":MBO22.objects.values_list('MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=request.user),"data7":request.user.email,"time":time,"ptotal":ptotal}
                    return render(request, "blog/Logout.html", context=params)
            if (time==2):
                task = get_object_or_404(MBO22, user=request.user)
                new_status = 1
                task.MBO22Q2 = new_status
                task.save()
            if (time==3):
                task = get_object_or_404(MBO22, user=request.user)
                new_status = 1
                task.MBO22Q3 = new_status
                task.save()
            if (time==4):
                task = get_object_or_404(MBO22, user=request.user)
                new_status = 1
                task.MBO22Q4 = new_status
                task.save()
            if (time!=5):
                time=0
                params = {"UserID":request.user,"data1":MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=request.user),"data2":MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=request.user),"data3":MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=request.user),"data4":MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=request.user),"data5":MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4').get(user=request.user),"data6":MBO22.objects.values_list('MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=request.user),"data7":request.user.email,"time":time}
                subject = request.user.username
                subject += " submeteru MBO, entre o sistema e aprove ou peça corrigir"
                message = request.user.username
                message += " submeteru MBO, entre o sistema e aprove ou peça corrigir"
                from_email = 'hyuma2331@gmail.com'  # 送信
#                from_email = 'sistema.rh@cosmotec.com.br'  # 送信
                recipient_list = [request.user.email]  # 宛先リスト
                send_mail(subject, message, from_email, recipient_list)
                return render(request, "blog/Logout.html", context=params)
    
    params = {
        "UserID":request.user,
        "data1":MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=request.user),
        "data2":MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=request.user),
        "data3":MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=request.user),
        "data4":MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=request.user),
        "data5":MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4').get(user=request.user),
        "data6":MBO22.objects.values_list('MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=request.user),
        "data7":request.user.email,
        "time":time,"time2":porta,"disp":disp,
        }
    return render(request, "blog/Logout.html",context=params)


@login_required
def ADC(request):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    porta=[]
    porta+=ADC22.objects.values_list('ADC22C').get(user=request.user)

    CG=RHDT.objects.values_list('ADC22G1').get(user=resultb6)
    CG1=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G2').get(user=resultb6)
    CG2=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G3').get(user=resultb6)
    CG3=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G4').get(user=resultb6)
    CG4=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G5').get(user=resultb6)
    CG5=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G6').get(user=resultb6)
    CG6=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G7').get(user=resultb6)
    CG7=str(CG[0])

    task = get_object_or_404(ADC22, user=request.user)

    time = int(porta[0])
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = xlsxwriter.Workbook(output)
            ws = book.add_worksheet('sua ADC')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'competencia geral', format)
            ws.write(1,2,'auto avaliação', format)
            ws.write(1,3,'avaliação por avaliador/a', format)
            ws.write(1,4,'observação por colaborador/a', format)
            ws.write(1,5,'observação por avaliador/a', format)
            ws.write(2,1,CG1, format)
            ws.write(3,1,CG2, format)
            ws.write(4,1,CG3, format)
            ws.write(5,1,CG4, format)
            ws.write(6,1,CG5, format)
            ws.write(7,1,CG6, format)
            ws.write(8,1,CG7, format)
            ws.write(2,2,task.ADC22G1C, format)
            ws.write(3,2,task.ADC22G2C, format)
            ws.write(4,2,task.ADC22G3C, format)
            ws.write(5,2,task.ADC22G4C, format)
            ws.write(6,2,task.ADC22G5C, format)
            ws.write(7,2,task.ADC22G6C, format)
            ws.write(8,2,task.ADC22G7C, format)
            ws.write(2,3,task.ADC22G1A, format)
            ws.write(3,3,task.ADC22G2A, format)
            ws.write(4,3,task.ADC22G3A, format)
            ws.write(5,3,task.ADC22G4A, format)
            ws.write(6,3,task.ADC22G5A, format)
            ws.write(7,3,task.ADC22G6A, format)
            ws.write(8,3,task.ADC22G7A, format)
            ws.write(2,4,task.ADC22G1OC, format)
            ws.write(3,4,task.ADC22G2OC, format)
            ws.write(4,4,task.ADC22G3OC, format)
            ws.write(5,4,task.ADC22G4OC, format)
            ws.write(6,4,task.ADC22G5OC, format)
            ws.write(7,4,task.ADC22G6OC, format)
            ws.write(8,4,task.ADC22G7OC, format)
            ws.write(2,5,task.ADC22G1O, format)
            ws.write(3,5,task.ADC22G2O, format)
            ws.write(4,5,task.ADC22G3O, format)
            ws.write(5,5,task.ADC22G4O, format)
            ws.write(6,5,task.ADC22G5O, format)
            ws.write(7,5,task.ADC22G6O, format)
            ws.write(8,5,task.ADC22G7O, format)
            ws.write(10,1,'competencia especifica', format)
            ws.write(10,2,'auto avaliação', format)
            ws.write(10,3,'avaliação por avaliador/a', format)
            ws.write(10,4,'observação por colaborador/a', format)
            ws.write(10,5,'observação por avaliador/a', format)
            ws.write(11,1,task.ADC22E1, format)
            ws.write(12,1,task.ADC22E2, format)
            ws.write(11,2,task.ADC22E1C, format)
            ws.write(12,2,task.ADC22E2C, format)
            ws.write(11,3,task.ADC22E1A, format)
            ws.write(12,3,task.ADC22E2A, format)
            ws.write(11,4,task.ADC22E1OC, format)
            ws.write(12,4,task.ADC22E2OC, format)
            ws.write(11,5,task.ADC22E1O, format)
            ws.write(12,5,task.ADC22E2O, format)
            ws.set_column('B:B',20)
            ws.set_column('C:D',15)
            ws.set_column('E:F',40)
            ws.set_row(2,30)
            ws.set_row(3,30)
            ws.set_row(4,30)
            ws.set_row(5,30)
            ws.set_row(6,30)
            ws.set_row(7,30)
            ws.set_row(8,30)
            ws.set_row(11,30)
            ws.set_row(12,30)
            book.close()
            output.seek(0)
            filename = 'sua ADC.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response
        elif "send" in request.POST:
            new_status = 1
            task.ADC22C = new_status
            task.save()
            time=1
            subject = request.user.username
            subject += " submeteru avaliação de competencia, entre o sistema e aprove ou peça corrigir"
            message = request.user.username
            message += " submeteru avaliação de competencia, entre o sistema e aprove ou peça corrigir"
#            from_email = 'hyuma2331@gmail.com'  # 送信
            from_email = 'sistema.rh@cosmotec.com.br'  # 送信
            recipient_list = [request.user.email]  # 宛先リスト
            send_mail(subject, message, from_email, recipient_list)


    CE1=request.user.first_name
    CE2=request.user.last_name

    task2 = get_object_or_404(ADC22, user=request.user)
    task2.ADC22E1 = CE1
    task2.ADC22E2 = CE2
    task2.save()

    Elist=[]  
    Elist+=RHDT.objects.values_list('ADC22E1C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E1D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12D').get(user=resultb6)


    CE3=''
    CE4=''

    for i in range(12):
        j=i*2
        k=j+1
        if str(Elist[j])==CE1:
            CE3=str(Elist[k])
        if str(Elist[j])==CE2:
            CE4=str(Elist[k])

    params = {"UserID":request.user,
        "data1":RHDT.objects.values_list('ADC22G1','ADC22G2','ADC22G3','ADC22G4','ADC22G5','ADC22G6','ADC22G7').get(user=resultb6),
        "data2":RHDT.objects.values_list('ADC22G1D','ADC22G2D','ADC22G3D','ADC22G4D','ADC22G5D','ADC22G6D','ADC22G7D').get(user=resultb6),
        "data3":ADC22.objects.values_list('ADC22G1C','ADC22G2C','ADC22G3C','ADC22G4C','ADC22G5C','ADC22G6C','ADC22G7C').get(user=request.user),
        "data4":ADC22.objects.values_list('ADC22G1A','ADC22G2A','ADC22G3A','ADC22G4A','ADC22G5A','ADC22G6A','ADC22G7A').get(user=request.user),
        "data3B":ADC22.objects.values_list('ADC22G1OC','ADC22G2OC','ADC22G3OC','ADC22G4OC','ADC22G5OC','ADC22G6OC','ADC22G7OC','ADC22E1OC','ADC22E2OC').get(user=request.user),
        "data5":ADC22.objects.values_list('ADC22G1O','ADC22G2O','ADC22G3O','ADC22G4O','ADC22G5O','ADC22G6O','ADC22G7O').get(user=request.user),
        "data6":request.user.first_name,
        "data7":request.user.last_name,
        "data81":CE3,
        "data82":CE4,
        "data9":ADC22.objects.values_list('ADC22E1C','ADC22E2C').get(user=request.user),
        "data10":ADC22.objects.values_list('ADC22E1A','ADC22E2A').get(user=request.user),
        "data11":ADC22.objects.values_list('ADC22E1O','ADC22E2O').get(user=request.user),
        "time":time,
        "avaliador":request.user.email
        }
    return render(request, "blog/ADC.html",context=params)

        
def ADC2(request,name):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    CG=RHDT.objects.values_list('ADC22G1').get(user=resultb6)
    CG1=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G2').get(user=resultb6)
    CG2=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G3').get(user=resultb6)
    CG3=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G4').get(user=resultb6)
    CG4=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G5').get(user=resultb6)
    CG5=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G6').get(user=resultb6)
    CG6=str(CG[0])
    CG=RHDT.objects.values_list('ADC22G7').get(user=resultb6)
    CG7=str(CG[0])

    colaborador = name#request.GET['name']

    result = ''
    result2 = ''
    result3 = ''
    result4 = ''
    result5 = ''
    result6 = 0
    result7 = ''
    result8 = ''
    result9 = ''
    result10 = ''
    result11 = ''
    result12 = ''
    result13 = ''
    result14 = ''
    result15 = ''
    object_list = User.objects.all()
    for i in object_list:
        result = i.username
        result2 += result
        result2 += ' '
        result3 = str(i.id)
        result4 += result3
        result4 += ' '
        result7 = i.email
        result8 += result7
        result8 += ' '
        result10 = i.first_name
        result11 += result10
        result11 += '^'
        result12 = i.last_name
        result13 += result12
        result13 += '^'
    result2=result2.split()
    result4=result4.split()
    result8=result8.split()
    result11=result11.split('^')
    result13=result13.split('^')
    k=int(len(result2))
    for j in range(k):
        if (str(result2[j])==colaborador):
            result6=int(result4[j])
            result9=str(result8[j])
            result14=str(result11[j])
            result15=str(result13[j])
    if result9 == str(request.user) :
        reject = 0
    else :
        reject = 1

    porta=[]
    porta+=ADC22.objects.values_list('ADC22C').get(user=result6)
    porta+=ADC22.objects.values_list('ADC22A').get(user=result6)

    if (int(porta[0])==0):
        time=0
    else :
        if (int(porta[1])==0):
            time=1
        else :
            time=2


    today = datetime.now()
    Y = today.year
    M = today.month
    D = today.day

    task = get_object_or_404(ADC22, user=result6)
    if (request.method == 'POST'):
        if "send" in request.POST:
            new_status = 1
            task.ADC22A = new_status
            task.ADC22Y = Y
            task.ADC22M = M
            task.ADC22D = D
            task.save()
            time=2
            subject = "sua avaliação de competencia foi aprovado, parabens!"
            message = "sua avaliação de competencia foi aprovado, parabens!"
            from_email = 'sistema.rh@cosmotec.com.br'  # 送信
            recipient_list = [colaborador]  # 宛先リスト
            send_mail(subject, message, from_email, recipient_list)
        elif "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = xlsxwriter.Workbook(output)
            ws = book.add_worksheet('avaliação dele(a)')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'competencia geral', format)
            ws.write(1,2,'auto avaliação', format)
            ws.write(1,3,'avaliação por avaliador/a', format)
            ws.write(1,4,'observação de colaborador/a', format)
            ws.write(1,5,'observação de avaliador/a', format)
            ws.write(2,1,CG1, format)
            ws.write(3,1,CG2, format)
            ws.write(4,1,CG3, format)
            ws.write(5,1,CG4, format)
            ws.write(6,1,CG5, format)
            ws.write(7,1,CG6, format)
            ws.write(8,1,CG7, format)
            ws.write(2,2,task.ADC22G1C, format)
            ws.write(3,2,task.ADC22G2C, format)
            ws.write(4,2,task.ADC22G3C, format)
            ws.write(5,2,task.ADC22G4C, format)
            ws.write(6,2,task.ADC22G5C, format)
            ws.write(7,2,task.ADC22G6C, format)
            ws.write(8,2,task.ADC22G7C, format)
            ws.write(2,3,task.ADC22G1A, format)
            ws.write(3,3,task.ADC22G2A, format)
            ws.write(4,3,task.ADC22G3A, format)
            ws.write(5,3,task.ADC22G4A, format)
            ws.write(6,3,task.ADC22G5A, format)
            ws.write(7,3,task.ADC22G6A, format)
            ws.write(8,3,task.ADC22G7A, format)
            ws.write(2,4,task.ADC22G1OC, format)
            ws.write(3,4,task.ADC22G2OC, format)
            ws.write(4,4,task.ADC22G3OC, format)
            ws.write(5,4,task.ADC22G4OC, format)
            ws.write(6,4,task.ADC22G5OC, format)
            ws.write(7,4,task.ADC22G6OC, format)
            ws.write(8,4,task.ADC22G7OC, format)
            ws.write(2,5,task.ADC22G1O, format)
            ws.write(3,5,task.ADC22G2O, format)
            ws.write(4,5,task.ADC22G3O, format)
            ws.write(5,5,task.ADC22G4O, format)
            ws.write(6,5,task.ADC22G5O, format)
            ws.write(7,5,task.ADC22G6O, format)
            ws.write(8,5,task.ADC22G7O, format)
            ws.write(10,1,'competencia geral', format)
            ws.write(10,2,'auto avaliação', format)
            ws.write(10,3,'avaliação por avaliador/a', format)
            ws.write(10,4,'observação de colaborador/a', format)
            ws.write(10,5,'observação de avaliador/a', format)
            ws.write(11,1,result14, format)
            ws.write(12,1,result15, format)
            ws.write(11,2,task.ADC22E1C, format)
            ws.write(12,2,task.ADC22E2C, format)
            ws.write(11,3,task.ADC22E1A, format)
            ws.write(12,3,task.ADC22E2A, format)
            ws.write(11,4,task.ADC22E1OC, format)
            ws.write(12,4,task.ADC22E2OC, format)
            ws.write(11,5,task.ADC22E1O, format)
            ws.write(12,5,task.ADC22E2O, format)
            ws.set_column('B:B',25)
            ws.set_column('C:D',15)
            ws.set_column('E:F',40)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(5,50)
            ws.set_row(6,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            ws.set_row(11,50)
            ws.set_row(12,50)
            book.close()
            output.seek(0)
            filename = 'avaliação dela(a).xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response


            

    params = {"UserID":request.user,
        "data1":RHDT.objects.values_list('ADC22G1','ADC22G2','ADC22G3','ADC22G4','ADC22G5','ADC22G6','ADC22G7').get(user=resultb6),
        "data2":ADC22.objects.values_list('ADC22G1C','ADC22G2C','ADC22G3C','ADC22G4C','ADC22G5C','ADC22G6C','ADC22G7C').get(user=result6),
        "data3":ADC22.objects.values_list('ADC22G1A','ADC22G2A','ADC22G3A','ADC22G4A','ADC22G5A','ADC22G6A','ADC22G7A').get(user=result6),
        "data4B":ADC22.objects.values_list('ADC22G1OC','ADC22G2OC','ADC22G3OC','ADC22G4OC','ADC22G5OC','ADC22G6OC','ADC22G7OC','ADC22E1OC','ADC22E2OC').get(user=result6),
        "data4":ADC22.objects.values_list('ADC22G1O','ADC22G2O','ADC22G3O','ADC22G4O','ADC22G5O','ADC22G6O','ADC22G7O').get(user=result6),
        "data5":ADC22.objects.values_list('ADC22E1','ADC22E2').get(user=result6),
        "data50":result14,
        "data51":result15,
        "data6":ADC22.objects.values_list('ADC22E1C','ADC22E2C').get(user=result6),
        "data7":ADC22.objects.values_list('ADC22E1A','ADC22E2A').get(user=result6),
        "data8":ADC22.objects.values_list('ADC22E1O','ADC22E2O').get(user=result6),
        "time":time,
        "avaliado":colaborador,
        "reject":reject,
        }
    return render(request, "blog/ADC2.html",context=params)


@login_required
def PDI(request):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    porta=[]
    porta+=PDI22.objects.values_list('PDI22C').get(user=request.user)
    time=int(porta[0])

    Glist=[]
#    Glist+=RHDT.objects.values_list('ADC22G1').get(user=resultb6),
    Glist.append('resultado')
    Glist+=ADC22.objects.values_list('ADC22G1C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G1A').get(user=request.user),
#    Glist+=RHDT.objects.values_list('ADC22G2').get(user=resultb6),
    Glist.append('cliente')
    Glist+=ADC22.objects.values_list('ADC22G2C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G2A').get(user=request.user),
#    Glist+=RHDT.objects.values_list('ADC22G3').get(user=resultb6),
    Glist.append('busca')
    Glist+=ADC22.objects.values_list('ADC22G3C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G3A').get(user=request.user),
#    Glist+=RHDT.objects.values_list('ADC22G4').get(user=resultb6),
    Glist.append('liderança')
    Glist+=ADC22.objects.values_list('ADC22G4C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G4A').get(user=request.user),
#    Glist+=RHDT.objects.values_list('ADC22G5').get(user=resultb6),
    Glist.append('senso')
    Glist+=ADC22.objects.values_list('ADC22G5C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G5A').get(user=request.user),
#    Glist+=RHDT.objects.values_list('ADC22G6').get(user=resultb6),
    Glist.append('comunicação')
    Glist+=ADC22.objects.values_list('ADC22G6C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G6A').get(user=request.user),
#    Glist+=RHDT.objects.values_list('ADC22G7').get(user=resultb6),
    Glist.append('relacionamento')
    Glist+=ADC22.objects.values_list('ADC22G7C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G7A').get(user=request.user),

    TC=ADC22.objects.values_list('ADC22G2C').get(user=request.user),
    TA=ADC22.objects.values_list('ADC22G2A').get(user=request.user),


    G1C=[]
    G2C=[]
    G3C=[]
    G1A=[]
    G2A=[]
    G3A=[]

    E1C=[]
    E2C=[]
    E1A=[]
    E2A=[]

    for i in range(7):
        j=i*3
        k=j+1
        l=k+1
        if str(Glist[j]) in str(PDI22.objects.values_list('PDI22G1C').get(user=request.user)).lower():
            G1C+=Glist[k]
            G1A+=Glist[l]
        if str(Glist[j]) in str(PDI22.objects.values_list('PDI22G2C').get(user=request.user)).lower():
            G2C+=Glist[k]
            G2A+=Glist[l]
        if str(Glist[j]) in str(PDI22.objects.values_list('PDI22G3C').get(user=request.user)).lower():
            G3C+=Glist[k]
            G3A+=Glist[l]

    E1C+=ADC22.objects.values_list('ADC22E1C').get(user=request.user)
    E1A+=ADC22.objects.values_list('ADC22E1A').get(user=request.user)
    E2C+=ADC22.objects.values_list('ADC22E2C').get(user=request.user)
    E2A+=ADC22.objects.values_list('ADC22E2A').get(user=request.user)

    CE1=request.user.first_name
    CE2=request.user.last_name

    task2 = get_object_or_404(PDI22, user=request.user)
    task2.PDI22E1C = CE1
    task2.PDI22E2C = CE2
    task2.save()

    if (G1C==[]):
        data11=''
    else :
        data11=G1C[0]
    if (G2C==[]):
        data21=''
    else :
        data21=G2C[0]
    if (G3C==[]):
        data31=''
    else :
        data31=G3C[0]
    if (G1A==[]):
        data12=''
    else :
        data12=G1A[0]
    if (G2A==[]):
        data22=''
    else :
        data22=G2A[0]
    if (G3A==[]):
        data32=''
    else :
        data32=G3A[0]

    task = get_object_or_404(PDI22, user=request.user)
    if (request.method == 'POST'):
        if "send" in request.POST:
            new_status = 1
            task.PDI22C = new_status
            task.save()
            time=1
            subject = request.user.username
            subject += " submeteru plano de desenvolvimento individual, entre o sistema e aprove ou peça corrigir"
            message = request.user.username
            message += " submeteru plano de desenvolvimento individual, entre o sistema e aprove ou peça corrigir"
#            from_email = 'hyuma2331@gmail.com'  # 送信
            from_email = 'sistema.rh@cosmotec.com.br'  # 送信
            recipient_list = [request.user.email]  # 宛先リスト
            send_mail(subject, message, from_email, recipient_list)
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = xlsxwriter.Workbook(output)
            ws = book.add_worksheet('seu PDI')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'foco de comeptencia geral', format)
            ws.write(1,2,'auto avaliação', format)
            ws.write(1,3,'avaliação por avaliador/a', format)
            ws.write(1,4,'ponto a desenvolver', format)
            ws.write(1,5,'Plano de Desenvolvimento Individual', format)
            ws.write(2,1,task.PDI22G1C, format)
            ws.write(2,2,data11, format)
            ws.write(2,3,data12, format)
            ws.write(2,4,task.PDI22G1PD, format)
            ws.write(2,5,task.PDI22G1, format)
            ws.write(3,1,task.PDI22G2C, format)
            ws.write(3,2,data21, format)
            ws.write(3,3,data22, format)
            ws.write(3,4,task.PDI22G2PD, format)
            ws.write(3,5,task.PDI22G2, format)
            ws.write(4,1,task.PDI22G3C, format)
            ws.write(4,2,data31, format)
            ws.write(4,3,data32, format)
            ws.write(4,4,task.PDI22G3PD, format)
            ws.write(4,5,task.PDI22G3, format)
            ws.write(6,1,'comeptencia especifica', format)
            ws.write(6,2,'auto avaliação', format)
            ws.write(6,3,'avaliação por avaliador/a', format)
            ws.write(6,4,'ponto a desenvolver', format)
            ws.write(6,5,'Plano de Desenvolvimento Individual', format)
            ws.write(7,1,task.PDI22E1C, format)
            ws.write(7,2,str(E1C[0]), format)
            ws.write(7,3,str(E1A[0]), format)
            ws.write(7,4,task.PDI22E1PD, format)
            ws.write(7,5,task.PDI22E1, format)
            ws.write(8,1,task.PDI22E2C, format)
            ws.write(8,2,str(E2C[0]), format)
            ws.write(8,3,str(E2A[0]), format)
            ws.write(8,4,task.PDI22E2PD, format)
            ws.write(8,5,task.PDI22E2, format)
            ws.set_column('B:B',20)
            ws.set_column('C:D',11)
            ws.set_column('E:F',33)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(5,50)
            ws.set_row(6,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            book.close()
            output.seek(0)
            filename = 'seu PDI.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response

    params = {"UserID":request.user,
        "data1":PDI22.objects.values_list('PDI22G1C','PDI22G1PD','PDI22G1').get(user=request.user),
        "data11":data11,
        "data12":data12,
        "data2":PDI22.objects.values_list('PDI22G2C','PDI22G2PD','PDI22G2').get(user=request.user),
        "data21":data21,
        "data22":data22,
        "data3":PDI22.objects.values_list('PDI22G3C','PDI22G3PD','PDI22G3').get(user=request.user),
        "data31":data31,
        "data32":data32,
        "data40":ADC22.objects.values_list('ADC22E1').get(user=request.user),
        "data4":PDI22.objects.values_list('PDI22E1PD','PDI22E1').get(user=request.user),
        "data41":E1C[0],
        "data42":E1A[0],
        "data50":ADC22.objects.values_list('ADC22E2').get(user=request.user),
        "data5":PDI22.objects.values_list('PDI22E2PD','PDI22E2').get(user=request.user),
        "data51":E2C[0],
        "data52":E2A[0],
        "time":time,
        "TC":TC,
        "TA":TA,
        "avaliador":request.user.email
        }
    return render(request, "blog/PDI.html",context=params)

def edit(request, num):
    obj = MBO22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = MBO22Q1Form(request.POST, instance=obj)
        friend.save()
        return redirect(to='/Logout')

    params = {
            'UserID': request.user,
            'data1': MBO22Q1Form(instance=obj),
            'data2': MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=request.user),
            'data3': MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=request.user),
            'data4': MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4','MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=request.user),
        }
    return render(request, 'blog/edit.html', params)

def edit2(request, num):
    obj = MBO22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = MBO22Q2Form(request.POST, instance=obj)
        friend.save()
        return redirect(to='/Logout')

    params = {
            'UserID': request.user,
            'title': 'alinha progresso com avaliador. depois de mexer, clique "salvar" em baixo senão vai perder recorde de mexer',
            'data1': MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=request.user),
            'data2': MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=request.user),
            'data3': MBO22Q2Form(instance=obj),
            'data4': MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=request.user),
            'data5': MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4','MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=request.user),
        }
    return render(request, 'blog/edit2.html', params)

def edit3(request, num):
    obj = MBO22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = MBO22Q3Form(request.POST, instance=obj)
        friend.save()
        return redirect(to='/Logout')

    params = {
            'UserID': request.user,
            'title': 'alinha progresso com avaliador. depois de mexer, clique "salvar" em baixo senão vai perder recorde de mexer',
            'data1': MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=request.user),
            'data2': MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=request.user),
            'data3': MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=request.user),
            'data4': MBO22Q3Form(instance=obj),
            'data5': MBO22.objects.values_list('MBO22A4','MBO22B4','MBO22C4','MBO22D4','MBO22E4','MBO22F4','MBO22G4','MBO22AR','MBO22BR','MBO22CR','MBO22DR','MBO22ER','MBO22FR','MBO22GR').get(user=request.user),
        }
    return render(request, 'blog/edit3.html', params)

def edit4(request, num):
    obj = MBO22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = MBO22Q4Form(request.POST, instance=obj)
        friend.save()
        return redirect(to='/Logout')

    params = {
            'UserID': request.user,
            'title': 'alinha progresso com avaliador. depois de mexer, clique "salvar" em baixo senão vai perder recorde de mexer',
            'data1': MBO22.objects.values_list('MBO22A1','MBO22B1','MBO22C1','MBO22D1','MBO22E1','MBO22F1','MBO22G1').get(user=request.user),
            'data2': MBO22.objects.values_list('MBO22AP','MBO22BP','MBO22CP','MBO22DP','MBO22EP','MBO22FP','MBO22GP').get(user=request.user),
            'data3': MBO22.objects.values_list('MBO22A2','MBO22B2','MBO22C2','MBO22D2','MBO22E2','MBO22F2','MBO22G2').get(user=request.user),
            'data4': MBO22.objects.values_list('MBO22A3','MBO22B3','MBO22C3','MBO22D3','MBO22E3','MBO22F3','MBO22G3').get(user=request.user),
            'data5': MBO22Q4Form(instance=obj),
        }
    return render(request, 'blog/edit4.html', params)


def editADC2(request, name):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])


    colaborador = name
    
    result = ''
    result2 = ''
    result3 = ''
    result4 = ''
    result5 = ''
    result6 = 0
    result7 = ''
    result8 = ''
    result9 = ''
    result10 = ''
    result11 = ''
    result12 = ''
    result13 = ''    
    object_list = User.objects.all()
    for i in object_list:
        result = i.username
        result2 += result
        result2 += '^'
        result3 = str(i.id)
        result4 += result3
        result4 += '^'
        result7 = i.first_name
        result8 += result7
        result8 += '^'
        result9 = i.last_name
        result10 += result9
        result10 += '^'
        result11 = i.email
        result12 += result11
        result12 += '^'
    result2=result2.split('^')
    result4=result4.split('^')
    result8=result8.split('^')
    result10=result10.split('^')
    result12=result12.split('^')
    k=int(len(result2))
    for j in range(k):
        if (str(result2[j])==colaborador):
            result6=int(result4[j])
            CE1=str(result8[j])
            CE3=str(result10[j])
            result13=str(result12[j])
    
    if result13 == str(request.user) :
        reject = 0
    else :
        reject = 1
    


    Elist=[]  
    Elist+=RHDT.objects.values_list('ADC22E1C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E1D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12D').get(user=resultb6)

    CE2=''
    CE4=''
    for i in range(12):
        j=i*2
        k=j+1
        if str(Elist[j])==CE1:
            CE2=str(Elist[k])
        if str(Elist[j])==CE3:
            CE4=str(Elist[k])

    time=1
    obj = ADC22.objects.get(user=result6)
    if (request.method == 'POST'):
        friend = ADC22AOForm(request.POST, instance=obj)
        friend.save()
        params = {"UserID":request.user,
            "data1":RHDT.objects.values_list('ADC22G1','ADC22G2','ADC22G3','ADC22G4','ADC22G5','ADC22G6','ADC22G7').get(user=resultb6),
            "data2":ADC22.objects.values_list('ADC22G1C','ADC22G2C','ADC22G3C','ADC22G4C','ADC22G5C','ADC22G6C','ADC22G7C').get(user=result6),
            "data3":ADC22.objects.values_list('ADC22G1A','ADC22G2A','ADC22G3A','ADC22G4A','ADC22G5A','ADC22G6A','ADC22G7A').get(user=result6),
            "data4B":ADC22.objects.values_list('ADC22G1OC','ADC22G2OC','ADC22G3OC','ADC22G4OC','ADC22G5OC','ADC22G6OC','ADC22G7OC','ADC22E1OC','ADC22E2OC').get(user=result6),
            "data4":ADC22.objects.values_list('ADC22G1O','ADC22G2O','ADC22G3O','ADC22G4O','ADC22G5O','ADC22G6O','ADC22G7O').get(user=result6),
            "data5":ADC22.objects.values_list('ADC22E1','ADC22E2').get(user=result6),
            "data6":ADC22.objects.values_list('ADC22E1C','ADC22E2C').get(user=result6),
            "data7":ADC22.objects.values_list('ADC22E1A','ADC22E2A').get(user=result6),
            "data8":ADC22.objects.values_list('ADC22E1O','ADC22E2O').get(user=result6),
            "time":time,
            "avaliado":colaborador,
            }
        return redirect('ADC2', name=colaborador)

    params = {
        'UserID':request.user,
        'form1': RHDT.objects.values_list('ADC22G1','ADC22G2','ADC22G3','ADC22G4','ADC22G5','ADC22G6','ADC22G7').get(user=resultb6),
        'form11': RHDT.objects.values_list('ADC22G1D','ADC22G2D','ADC22G3D','ADC22G4D','ADC22G5D','ADC22G6D','ADC22G7D').get(user=resultb6),
        'CE1' : CE1,
        'CE2' : CE2,
        'CE3' : CE3,
        'CE4' : CE4,
        'form2': ADC22.objects.values_list('ADC22G1C','ADC22G2C','ADC22G3C','ADC22G4C','ADC22G5C','ADC22G6C','ADC22G7C','ADC22E1C','ADC22E2C').get(user=result6),
        'form2B': ADC22.objects.values_list('ADC22G1OC','ADC22G2OC','ADC22G3OC','ADC22G4OC','ADC22G5OC','ADC22G6OC','ADC22G7OC','ADC22E1OC','ADC22E2OC').get(user=result6),
        'form3': ADC22AOForm(instance=obj),
        'avaliado' : colaborador,
        'reject' : reject,
    }
    return render(request, 'blog/editADC2.html', params)


def editADC(request, num):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])


    Elist=[]  
    Elist+=RHDT.objects.values_list('ADC22E1C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E1D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12D').get(user=resultb6)

    CE1=request.user.first_name
    CE2=''
    CE3=request.user.last_name
    CE4=''

    for i in range(12):
        j=i*2
        k=j+1
        if str(Elist[j])==CE1:
            CE2=str(Elist[k])
        if str(Elist[j])==CE3:
            CE4=str(Elist[k])


    obj = ADC22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = ADC22CForm(request.POST, instance=obj)
        friend.save()
        return redirect(to='/ADC')

    params = {
        'UserID':request.user,
        'form1': RHDT.objects.values_list('ADC22G1','ADC22G2','ADC22G3','ADC22G4','ADC22G5','ADC22G6','ADC22G7').get(user=resultb6),
        'form11': RHDT.objects.values_list('ADC22G1D','ADC22G2D','ADC22G3D','ADC22G4D','ADC22G5D','ADC22G6D','ADC22G7D').get(user=resultb6),
        'form51': CE1,
        'form52': CE2,
        'form53': CE3,
        'form54': CE4,
        'form2': ADC22CForm(instance=obj),
        'form3': ADC22.objects.values_list('ADC22G1A','ADC22G2A','ADC22G3A','ADC22G4A','ADC22G5A','ADC22G6A','ADC22G7A','ADC22E1A','ADC22E2A').get(user=request.user),
        'form4': ADC22.objects.values_list('ADC22G1O','ADC22G2O','ADC22G3O','ADC22G4O','ADC22G5O','ADC22G6O','ADC22G7O','ADC22E1O','ADC22E2O').get(user=request.user),
    }
    return render(request, 'blog/editADC.html', params)

def editPDI(request, num):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    Glist=[]
    Glist+=RHDT.objects.values_list('ADC22G1').get(user=resultb6),
    Glist+=ADC22.objects.values_list('ADC22G1C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G1A').get(user=request.user),
    Glist+=RHDT.objects.values_list('ADC22G2').get(user=resultb6),
    Glist+=ADC22.objects.values_list('ADC22G2C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G2A').get(user=request.user),
    Glist+=RHDT.objects.values_list('ADC22G3').get(user=resultb6),
    Glist+=ADC22.objects.values_list('ADC22G3C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G3A').get(user=request.user),
    Glist+=RHDT.objects.values_list('ADC22G4').get(user=resultb6),
    Glist+=ADC22.objects.values_list('ADC22G4C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G4A').get(user=request.user),
    Glist+=RHDT.objects.values_list('ADC22G5').get(user=resultb6),
    Glist+=ADC22.objects.values_list('ADC22G5C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G5A').get(user=request.user),
    Glist+=RHDT.objects.values_list('ADC22G6').get(user=resultb6),
    Glist+=ADC22.objects.values_list('ADC22G6C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G6A').get(user=request.user),
    Glist+=RHDT.objects.values_list('ADC22G7').get(user=resultb6),
    Glist+=ADC22.objects.values_list('ADC22G7C').get(user=request.user),
    Glist+=ADC22.objects.values_list('ADC22G7A').get(user=request.user),

    G1C=['']
    G2C=['']
    G3C=['']
    G1A=['']
    G2A=['']
    G3A=['']

    E1C=['']
    E2C=['']
    E1A=['']
    E2A=['']

    for i in range(7):
        j=i*3
        k=j+1
        l=k+1
        if (PDI22.objects.values_list('PDI22G1C').get(user=request.user)==Glist[j]):
            G1C=Glist[k]
            G1A=Glist[l]
        if (PDI22.objects.values_list('PDI22G2C').get(user=request.user)==Glist[j]):
            G2C=Glist[k]
            G2A=Glist[l]
        if (PDI22.objects.values_list('PDI22G3C').get(user=request.user)==Glist[j]):
            G3C=Glist[k]
            G3A=Glist[l]

    E1C+=ADC22.objects.values_list('ADC22E1C').get(user=request.user)
    E1A+=ADC22.objects.values_list('ADC22E1A').get(user=request.user)
    E2C+=ADC22.objects.values_list('ADC22E2C').get(user=request.user)
    E2A+=ADC22.objects.values_list('ADC22E2A').get(user=request.user)

    obj = PDI22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = PDI22Form(request.POST, instance=obj)
        friend.save()
        return redirect(to='/PDI')

    params = {
        'UserID':request.user,
        'form1': PDI22Form(instance=obj),
        'form2': request.user.first_name,
        'form3': request.user.last_name,
        'G1C':G1C[0],'G2C':G2C[0],'G3C':G3C[0],'G1A':G1A[0],'G2A':G2A[0],'G3A':G3A[0],'E1C':E1C[0],'E2C':E2C[0],'E1A':E1A[0],'E2A':E2A[0],
    }
    return render(request, 'blog/editPDI.html', params)


#ホーム
@login_required
def home(request):


    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    MBO23TIME=[]
    MBO23TIME+=RHDT.objects.values_list('MBOTIME23').get(user=resultb6)

    MBO24TIME=[]
    MBO24TIME+=RHDT.objects.values_list('MBOTIME24').get(user=resultb6)

    ADC23TIME=[]
    ADC23TIME+=RHDT.objects.values_list('ADCTIME23').get(user=resultb6)

    ADC24TIME=[]
    ADC24TIME+=RHDT.objects.values_list('ADCTIME24').get(user=resultb6)

    obj = MBO22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = DeptForm(request.POST, instance=obj)
        friend.save()

    dept=[]
    dept+=MBO22.objects.values_list('DEPT').get(user=request.user)
    dept1=str(dept[0])
    if dept1 == '':
        time=0
    else:
        time=1

    equipe = 0
    object_list = User.objects.all()
    for i in object_list:
        if str(i.email) == str(request.user):
            equipe =1
     
    params = {
        "UserID":request.user,
        "MBO23TIME" : int(MBO23TIME[0]),
        "MBO24TIME" : int(MBO24TIME[0]),
        "ADC23TIME" : int(ADC23TIME[0]),
        "ADC24TIME" : int(ADC24TIME[0]),
        "time" : time,
        "dept": DeptForm(instance=obj),
        "equipe" : equipe,
        }
    return render(request, "blog/home.html",context=params)


@login_required
def RH(request):
    params = {"UserID":request.user,}
    return render(request, "blog/RH.html",context=params)

@login_required
def ES(request):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])


    Elist=[]  
    Elist+=RHDT.objects.values_list('ADC22E1C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E1D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11D').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12D').get(user=resultb6)


    ES='0,'
    AA=[]
    object_list = User.objects.all()
#    for k in range(j):
    for i in object_list:
        ES+=i.username
        ES+=','
        BB=[]
        try:
            BB+=MBO22.objects.values_list('DEPT').get(user=i.id)
            ES+=str(BB[0])
            ES+=','
        except:
            ES+='departamento não definido'
            ES+=','
        ES+=i.email
        ES+=','
        CE3=0
        for m in object_list:
            if i.email==m.username:
                ES+='OK'
                ES+=','
                CE3=1
        if CE3==0:
            ES+='não'
            ES+=','
        ES+=i.first_name
        ES+=','
        CE1=0
        CE2=0
        for j in range(12):
            k=j*2
            if str(Elist[k])!='':
                if str(Elist[k])==i.first_name:
                    ES+='OK'
                    ES+=','
                    CE1=1
        if CE1==0:
            ES+='não'
            ES+=','
        ES+=i.last_name
        ES+=','
        for j in range(12):
            k=j*2
            if str(Elist[k])!='':
                if str(Elist[k])==i.last_name:
                    ES+='OK'
                    ES+=','
                    CE2=1
        if CE2==0:
            ES+='não'
            ES+=','
        try :
            AA+=MBO22.objects.values_list('MBO22Q1Y').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q1M').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q1D').get(user=i.id)
            if ((int(AA[0])>0)):
                ES+=str(AA[2])+ '-'+str(AA[1])+'-'+str(AA[0])
            else:
                ES+='not yet'
        except:
            ES+='sem MBO'
        ES+=','
        AA=[]
        try:
            AA+=MBO22.objects.values_list('MBO22Q2Y').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q2M').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q2D').get(user=i.id)
            if ((int(AA[0])>0)):
                ES+=str(AA[2])+ '-'+str(AA[1])+'-'+str(AA[0])
            else:
                ES+='not yet'
        except:
            ES+='sem MBO'
        ES+=','
        AA=[]
        try :
            AA+=MBO22.objects.values_list('MBO22Q3Y').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q3M').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q3D').get(user=i.id)
            if ((int(AA[0])>0)):
                ES+=str(AA[2])+ '-'+str(AA[1])+'-'+str(AA[0])
            else:
                ES+='not yet'
        except:
            ES+='sem MBO'
        ES+=','
        AA=[]
        try:
            AA+=MBO22.objects.values_list('MBO22Q4Y').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q4M').get(user=i.id)
            AA+=MBO22.objects.values_list('MBO22Q4D').get(user=i.id)
            if ((int(AA[0])>0)):
                ES+=str(AA[2])+ '-'+str(AA[1])+'-'+str(AA[0])
            else:
                ES+='not yet'
        except:
            ES+='sem MBO'
        ES+=','
        AA=[]
        try:
            AA+=ADC22.objects.values_list('ADC22Y').get(user=i.id)
            AA+=ADC22.objects.values_list('ADC22M').get(user=i.id)
            AA+=ADC22.objects.values_list('ADC22D').get(user=i.id)
            if ((int(AA[0])>0)):
                ES+=str(AA[2])+ '-'+str(AA[1])+'-'+str(AA[0])
            else:
                ES+='not yet'
        except:
            ES+='sem ADC'
        ES+=','
        AA=[]
        try:
            AA+=PDI22.objects.values_list('PDI22Y').get(user=i.id)
            AA+=PDI22.objects.values_list('PDI22M').get(user=i.id)
            AA+=PDI22.objects.values_list('PDI22D').get(user=i.id)
            if ((int(AA[0])>0)):
                ES+=str(AA[2])+ '-'+str(AA[1])+'-'+str(AA[0])
            else:
                ES+='not yet'
        except:
            ES+='sem PDI'
        ES+=','
        AA=[]
        ES+='#'
        ES+=','
    ES=ES.split(',')
    del ES[-1]
    #j=int(int(len(ES))/k)
    #ES=np.reshape(ES,(j,k))
    #ES=pd.DataFrame(ES)
    params = {"UserID":request.user,"time":ES}
    return render(request, "blog/ES.html",context=params)


@login_required
def MBO22RH(request):
    
    object_list = User.objects.all()
    AA=''
    BB=[]
    for i in object_list:
        AA+='#'
        AA+='^'
        AA+=i.username
        AA+='^'
        AA+='#'
        AA+='^'
        try :
            BB=MBO22.objects.values_list('DEPT').get(user=i.id)
            AA+=BB[0]
            BB=[]
        except :
            AA+='dept não definido'
        AA+='^'
        AA+='#'
        AA+='^'
        try:
            BB=MBO22.objects.values_list('MBO22A1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22AP').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22A2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22A3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22A4').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22AR').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=MBO22.objects.values_list('MBO22B1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22BP').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22B2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22B3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22B4').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22BR').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=MBO22.objects.values_list('MBO22C1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22CP').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22C2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22C3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22C4').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22CR').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=MBO22.objects.values_list('MBO22D1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22DP').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22D2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22D3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22D4').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22DR').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=MBO22.objects.values_list('MBO22E1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22EP').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22E2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22E3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22E4').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22ER').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=MBO22.objects.values_list('MBO22F1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22FP').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22F2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22F3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22F4').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22FR').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=MBO22.objects.values_list('MBO22G1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22GP').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22G2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22G3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22G4').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=MBO22.objects.values_list('MBO22GR').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            peso1=MBO22.objects.values_list('MBO22AP').get(user=i.id)
            res1=MBO22.objects.values_list('MBO22AR').get(user=i.id)
            peso2=MBO22.objects.values_list('MBO22BP').get(user=i.id)
            res2=MBO22.objects.values_list('MBO22BR').get(user=i.id)
            peso3=MBO22.objects.values_list('MBO22CP').get(user=i.id)
            res3=MBO22.objects.values_list('MBO22CR').get(user=i.id)
            peso4=MBO22.objects.values_list('MBO22DP').get(user=i.id)
            res4=MBO22.objects.values_list('MBO22DR').get(user=i.id)
            peso5=MBO22.objects.values_list('MBO22EP').get(user=i.id)
            res5=MBO22.objects.values_list('MBO22ER').get(user=i.id)
            peso6=MBO22.objects.values_list('MBO22FP').get(user=i.id)
            res6=MBO22.objects.values_list('MBO22FR').get(user=i.id)
            peso7=MBO22.objects.values_list('MBO22GP').get(user=i.id)
            res7=MBO22.objects.values_list('MBO22GR').get(user=i.id)
            pefo=int(peso1[0])*int(res1[0])+int(peso2[0])*int(res2[0])+int(peso3[0])*int(res3[0])+int(peso4[0])*int(res4[0])+int(peso5[0])*int(res5[0])+int(peso6[0])*int(res6[0])+int(peso7[0])*int(res7[0])
            pefo=pefo/100
            AA+=str(pefo)
            AA+='^'
            AA+='#'
            AA+='^'
        except:
            AA+='não tem MBO'
            AA+='^'
    AA=AA.split('^')
    del AA[-1]
    params = {"UserID":request.user,"time":AA}
    return render(request, "blog/MBO22RH.html",context=params)

@login_required
def ADC22RHA(request):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])


    object_list = User.objects.all()
    AA=''
    BB=[]
    for i in object_list:
        AA+='#'
        AA+='^'
        AA+=i.username
        AA+='^'
        AA+='#'
        AA+='^'
        try :
            BB=MBO22.objects.values_list('DEPT').get(user=i.id)
            AA+=BB[0]
            BB=[]
        except :
            AA+='dept não definido'
        AA+='^'
        AA+='#'
        AA+='^'
        try:
            CC=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
            BB=RHDT.objects.values_list('ADC22G1').get(user=resultb6)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G1OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G1O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=RHDT.objects.values_list('ADC22G2').get(user=resultb6)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G2OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G2O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=RHDT.objects.values_list('ADC22G3').get(user=resultb6)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G3OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G3O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=RHDT.objects.values_list('ADC22G4').get(user=resultb6)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G4OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G4O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=RHDT.objects.values_list('ADC22G5').get(user=resultb6)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G5OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G5O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=RHDT.objects.values_list('ADC22G6').get(user=resultb6)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G6OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G6O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=RHDT.objects.values_list('ADC22G7').get(user=resultb6)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G7OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22G7O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            AA+=i.first_name
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E1OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E1O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            AA+=i.last_name
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E2OC').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E2O').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
        except:
            AA+='não tem avaliação de competencia'
            AA+='^'
    AA=AA.split('^')
    del AA[-1]
    params = {"UserID":request.user,"time":AA}
    return render(request, "blog/ADC22RHA.html",context=params)

@login_required
def PDI22RH(request):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])


    object_list = User.objects.all()
    AA=''
    BB=[]
    for i in object_list:
        AA+='#'
        AA+='^'
        AA+=i.username
        AA+='^'
        AA+='#'
        AA+='^'
        try :
            BB=MBO22.objects.values_list('DEPT').get(user=i.id)
            AA+=BB[0]
            BB=[]
        except :
            AA+='dept não definido'
        AA+='^'
        AA+='#'
        AA+='^'      

        try:
            Glist=[]
            Glist.append('resultado')   
            Glist.append('cliente')   
            Glist.append('busca')   
            Glist.append('liderança')   
            Glist.append('senso')   
            Glist.append('comunicação')   
            Glist.append('relacionamento')   
            Glist.append(ADC22.objects.values_list('ADC22G1C').get(user=i.id)),
            Glist+=ADC22.objects.values_list('ADC22G2C').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G3C').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G4C').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G5C').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G6C').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G7C').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G1A').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G2A').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G3A').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G4A').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G5A').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G6A').get(user=i.id),
            Glist+=ADC22.objects.values_list('ADC22G7A').get(user=i.id),

            BB=PDI22.objects.values_list('PDI22G1C').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            for xx in range(7):
                yy=xx+7
                zz=xx+14
                if str(Glist[xx]) in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
                    AA+=str(Glist[yy])
                    AA+='^'
                    AA+=str(Glist[zz])
                    AA+='^'
            BB=PDI22.objects.values_list('PDI22G1PD').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=PDI22.objects.values_list('PDI22G1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=PDI22.objects.values_list('PDI22G2C').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            for xx in range(7):
                yy=xx+7
                zz=xx+14
                if str(Glist[xx]) in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                    AA+=str(Glist[yy])
                    AA+='^'
                    AA+=str(Glist[zz])
                    AA+='^'
            BB=PDI22.objects.values_list('PDI22G2PD').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=PDI22.objects.values_list('PDI22G2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            BB=PDI22.objects.values_list('PDI22G3C').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            for xx in range(7):
                yy=xx+7
                zz=xx+14
                if str(Glist[xx]) in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                    AA+=str(Glist[yy])
                    AA+='^'
                    AA+=str(Glist[zz])
                    AA+='^'
            BB=PDI22.objects.values_list('PDI22G3PD').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=PDI22.objects.values_list('PDI22G3').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            #BB=PDI22.objects.values_list('PDI22E1C').get(user=i.id)
            AA+=i.first_name#BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=PDI22.objects.values_list('PDI22E1PD').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=PDI22.objects.values_list('PDI22E1').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
            #BB=PDI22.objects.values_list('PDI22E2C').get(user=i.id)
            AA+=i.last_name#BB[0]
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
            AA+=str(BB[0])
            BB=[]
            AA+='^'
            BB=PDI22.objects.values_list('PDI22E2PD').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            BB=PDI22.objects.values_list('PDI22E2').get(user=i.id)
            AA+=BB[0]
            BB=[]
            AA+='^'
            AA+='#'
            AA+='^'
        except:
            AA+='não tem PDI'
            AA+='^'
    AA=AA.split('^')
    del AA[-1]
    params = {"UserID":request.user,"time":AA}
    return render(request, "blog/PDI22RH.html",context=params)


def MBO22DL(request):
    
    A1=0
    A2=0
    A3=0
    A4=0
    A5=0
    A6=0
    A7=0
    A8=0
    A9=0
    A10=0
    A11=0
    A12=0
    B1=0
    B2=0
    B3=0
    B4=0
    B5=0
    B6=0
    B7=0
    B8=0
    B9=0
    B10=0
    B11=0
    B12=0
    C1=0
    C2=0
    C3=0
    C4=0
    C5=0
    C6=0
    C7=0
    C8=0
    C9=0
    C10=0
    C11=0
    C12=0
    D1=0
    D2=0
    D3=0
    D4=0
    D5=0
    D6=0
    D7=0
    D8=0
    D9=0
    D10=0
    D11=0
    D12=0

    NA1=''
    NA2=''
    NA3=''
    NA4=''
    NA5=''
    NA6=''
    NA7=''
    NA8=''
    NA9=''
    NA10=''
    NA11=''
    NA12=''
    NB1=''
    NB2=''
    NB3=''
    NB4=''
    NB5=''
    NB6=''
    NB7=''
    NB8=''
    NB9=''
    NB10=''
    NB11=''
    NB12=''
    NC1=''
    NC2=''
    NC3=''
    NC4=''
    NC5=''
    NC6=''
    NC7=''
    NC8=''
    NC9=''
    NC10=''
    NC11=''
    NC12=''
    ND1=''
    ND2=''
    ND3=''
    ND4=''
    ND5=''
    ND6=''
    ND7=''
    ND8=''
    ND9=''
    ND10=''
    ND11=''
    ND12=''

    test=0
    BB1=''
    BB2=''
    BB3=''
    BB4=''
    CC=''

    object_list = User.objects.all()
#    for k in range(j):
    for i in object_list:
        test=[]
        try :
            test+=MBO22.objects.values_list('MBO22Q1M').get(user=i.id)
            if (test[0]==0):
                BB1+=i.username
                BB1+=','
            if (test[0]==1):
                A1+=1
                NA1+=i.username
                NA1+='\n'
            if (test[0]==2):
                A2+=1
                NA2+=i.username
                NA2+='\n'
            if (test[0]==3):
                A3+=1
                NA3+=i.username
                NA3+='\n'
            if (test[0]==4):
                A4+=1
                NA4+=i.username
                NA4+='\n'
            if (test[0]==5):
                A5+=1
                NA5+=i.username
                NA5+='\n'
            if (test[0]==6):
                A6+=1
                NA6+=i.username
                NA6+='\n'
            if (test[0]==7):
                A7+=1
                NA7+=i.username
                NA7+='\n'
            if (test[0]==8):
                A8+=1
                NA8+=i.username
                NA8+='\n'
            if (test[0]==9):
                A9+=1
                NA19+=i.username
                NA19='\n'
            if (test[0]==10):
                A10+=1
                NA10+=i.username
                NA10+='\n'
            if (test[0]==11):
                A11+=1
                NA11+=i.username
                NA11+='\n'
            if (test[0]==12):
                A12+=1
                NA12+=i.username
                NA12+='\n'
            test=[]
            test+=MBO22.objects.values_list('MBO22Q2M').get(user=i.id)
            if (test[0]==0):
                BB2+=i.username
                BB2+=','
            if (test[0]==1):
                B1+=1
                NB1+=i.username
                NB1+='\n'
            if (test[0]==2):
                B2+=1
                NB2+=i.username
                NB2+='\n'
            if (test[0]==3):
                B3+=1
                NB3+=i.username
                NB3+='\n'
            if (test[0]==4):
                B4+=1
                NB4+=i.username
                NB4+='\n'
            if (test[0]==5):
                B5+=1
                NB5+=i.username
                NB5+='\n'
            if (test[0]==6):
                B6+=1
                NB6+=i.username
                NB6+='\n'
            if (test[0]==7):
                B7+=1
                NB7+=i.username
                NB7+='\n'
            if (test[0]==8):
                B8+=1
                NB8+=i.username
                NB8+='\n'
            if (test[0]==9):
                B9+=1
                NB9+=i.username
                NB9+='\n'
            if (test[0]==10):
                B10+=1
                NB10+=i.username
                NB10+='\n'
            if (test[0]==11):
                B11+=1
                NB11+=i.username
                NB11+='\n'
            if (test[0]==12):
                B12+=1
                NB12+=i.username
                NB12+='\n'
            test=[]
            test+=MBO22.objects.values_list('MBO22Q3M').get(user=i.id)
            if (test[0]==0):
                BB3+=i.username
                BB3+=','
            if (test[0]==1):
                C1+=1
                NC1+=i.username
                NC1+='\n'
            if (test[0]==2):
                C2+=1
                NC2+=i.username
                NC2+='\n'
            if (test[0]==3):
                C3+=1
                NC3+=i.username
                NC3+='\n'
            if (test[0]==4):
                C4+=1
                NC4+=i.username
                NC4+='\n'
            if (test[0]==5):
                C5+=1
                NC5+=i.username
                NC5+='\n'
            if (test[0]==6):
                C6+=1
                NC6+=i.username
                NC6+='\n'
            if (test[0]==7):
                C7+=1
                NC7+=i.username
                NC7+='\n'
            if (test[0]==8):
                C8+=1
                NC8+=i.username
                NC8+='\n'
            if (test[0]==9):
                C9+=1
                NC9+=i.username
                NC9+='\n'
            if (test[0]==10):
                C10+=1
                NC10+=i.username
                NC10+='\n'
            if (test[0]==11):
                C11+=1
                NC11+=i.username
                NC11+='\n'
            if (test[0]==12):
                C12+=1
                NC12+=i.username
                NC12+='\n'
            test=[]
            test+=MBO22.objects.values_list('MBO22Q4M').get(user=i.id)
            if (test[0]==0):
                BB4+=i.username
                BB4+=','
            if (test[0]==1):
                D1+=1
                ND1+=i.username
                ND1+='\n'
            if (test[0]==2):
                D2+=1
                ND2+=i.username
                ND2+='\n'
            if (test[0]==3):
                D3+=1
                ND3+=i.username
                ND3+='\n'
            if (test[0]==4):
                D4+=1
                ND4+=i.username
                ND4+='\n'
            if (test[0]==5):
                D5+=1
                ND5+=i.username
                ND5+='\n'
            if (test[0]==6):
                D6+=1
                ND6+=i.username
                ND6+='\n'
            if (test[0]==7):
                D7+=1
                ND7+=i.username
                ND7+='\n'
            if (test[0]==8):
                D8+=1
                ND8+=i.username
                ND8+='\n'
            if (test[0]==9):
                D9+=1
                ND9+=i.username
                ND9+='\n'
            if (test[0]==10):
                D10+=1
                ND10+=i.username
                ND10+='\n'
            if (test[0]==11):
                D11+=1
                ND11+=i.username
                ND11+='\n'
            if (test[0]==12):
                D12+=1
                ND12+=i.username
                ND12+='\n'
        except:
            CC+=i.username
            CC+=','
    BB1=BB1.split(',')
    BB2=BB2.split(',')
    BB3=BB3.split(',')
    BB4=BB4.split(',')
    CC=CC.split(',')
    AT=A1+A2+A3+A4+A5+A6+A7+A8+A9+A10+A11+A12
    BT=B1+B2+B3+B4+B5+B6+B7+B8+B9+B10+B11+B12
    CT=C1+C2+C3+C4+C5+C6+C7+C8+C9+C10+C11+C12
    DT=D1+D2+D3+D4+D5+D6+D7+D8+D9+D10+D11+D12
    params = {
        "UserID":request.user,
        "A1":A1,"A2":A2,"A3":A3,"A4":A4,"A5":A5,"A6":A6,"A7":A7,"A8":A8,"A9":A9,"A10":A10,"A11":A11,"A12":A12,
        "B1":B1,"B2":B2,"B3":B3,"B4":B4,"B5":B5,"B6":B6,"B7":B7,"B8":B8,"B9":B9,"B10":B10,"B11":B11,"B12":B12,
        "C1":C1,"C2":C2,"C3":C3,"C4":C4,"C5":C5,"C6":C6,"C7":C7,"C8":C8,"C9":C9,"C10":C10,"C11":C11,"C12":C12,
        "D1":D1,"D2":D2,"D3":D3,"D4":D4,"D5":D5,"D6":D6,"D7":D7,"D8":D8,"D9":D9,"D10":D10,"D11":D11,"D12":D12,
        "NA1":NA1,"NA2":NA2,"NA3":NA3,"NA4":NA4,"NA5":NA5,"NA6":NA6,"NA7":NA7,"NA8":NA8,"NA9":NA9,"NA10":NA10,"NA11":NA11,"NA12":NA12,
        "NB1":NB1,"NB2":NB2,"NB3":NB3,"NB4":NB4,"NB5":NB5,"NB6":NB6,"NB7":NB7,"NB8":NB8,"NB9":NB9,"NB10":NB10,"NB11":NB11,"NB12":NB12,
        "NC1":NC1,"NC2":NC2,"NC3":NC3,"NC4":NC4,"NC5":NC5,"NC6":NC6,"NC7":NC7,"NC8":NC8,"NC9":NC9,"NC10":NC10,"NC11":NC11,"NC12":NC12,
        "ND1":ND1,"ND2":ND2,"ND3":ND3,"ND4":ND4,"ND5":ND5,"ND6":ND6,"ND7":ND7,"ND8":ND8,"ND9":ND9,"ND10":ND10,"ND11":ND11,"ND12":ND12,
        "BB1":BB1,"BB2":BB2,"BB3":BB3,"BB4":BB4,
        "CC":CC,
        "AT":AT,"BT":BT,"CT":CT,"DT":DT,
        }
    return render(request, "blog/MBO22DL.html",context=params)


def ADC22DL(request):
    
    A1=0
    A2=0
    A3=0
    A4=0
    A5=0
    A6=0
    A7=0
    A8=0
    A9=0
    A10=0
    A11=0
    A12=0

    B1=0
    B2=0
    B3=0
    B4=0
    B5=0
    B6=0
    B7=0
    B8=0
    B9=0
    B10=0
    B11=0
    B12=0

    NA1=''
    NA2=''
    NA3=''
    NA4=''
    NA5=''
    NA6=''
    NA7=''
    NA8=''
    NA9=''
    NA10=''
    NA11=''
    NA12=''

    NB1=''
    NB2=''
    NB3=''
    NB4=''
    NB5=''
    NB6=''
    NB7=''
    NB8=''
    NB9=''
    NB10=''
    NB11=''
    NB12=''

    test=0
    BB1=''
    BB2=''
    CC1=''
    CC2=''
    object_list = User.objects.all()
#    for k in range(j):
    for i in object_list:
        test=[]
        try :
            test+=ADC22.objects.values_list('ADC22M').get(user=i.id)
            if (test[0]==0):
                BB1+=i.username
                BB1+=','
            if (test[0]==1):
                A1+=1
                NA1+=i.username
                NA1+='\n'
            if (test[0]==2):
                A2+=1
                NA2+=i.username
                NA2+='\n'
            if (test[0]==3):
                A3+=1
                NA3+=i.username
                NA3+='\n'
            if (test[0]==4):
                A4+=1
                NA4+=i.username
                NA4+='\n'
            if (test[0]==5):
                A5+=1
                NA5+=i.username
                NA5+='\n'
            if (test[0]==6):
                A6+=1
                NA6+=i.username
                NA6+='\n'
            if (test[0]==7):
                A7+=1
                NA7+=i.username
                NA7+='\n'
            if (test[0]==8):
                A8+=1
                NA8+=i.username
                NA8+='\n'
            if (test[0]==9):
                A9+=1
                NA9+=i.username
                NA9+='\n'
            if (test[0]==10):
                A10+=1
                NA10+=i.username
                NA10+='\n'
            if (test[0]==11):
                A11+=1
                NA11+=i.username
                NA11+='\n'
            if (test[0]==12):
                A12+=1
                NA12+=i.username
                NA12+='\n'
        except:
            CC1+=i.username
            CC1+=','
        test=[]
        try :
            test+=PDI22.objects.values_list('PDI22M').get(user=i.id)
            if (test[0]==0):
                BB2+=i.username
                BB2+=','
            if (test[0]==1):
                B1+=1
                NB1+=i.username
                NB1+='\n'
            if (test[0]==2):
                B2+=1
                NB2+=i.username
                NB2+='\n'
            if (test[0]==3):
                B3+=1
                NB3+=i.username
                NB3+='\n'
            if (test[0]==4):
                B4+=1
                NB4+=i.username
                NB4+='\n'
            if (test[0]==5):
                B5+=1
                NB5+=i.username
                NB5+='\n'
            if (test[0]==6):
                B6+=1
                NB6+=i.username
                NB6+='\n'
            if (test[0]==7):
                B7+=1
                NB7+=i.username
                NB7+='\n'
            if (test[0]==8):
                B8+=1
                NB8+=i.username
                NB8+='\n'
            if (test[0]==9):
                B9+=1
                NB9+=i.username
                NB9+='\n'
            if (test[0]==10):
                B10+=1
                NB10+=i.username
                NB10+='\n'
            if (test[0]==11):
                B11+=1
                NB11+=i.username
                NB11+='\n'
            if (test[0]==12):
                B12+=1
                NB12+=i.username
                NB12+='\n'
        except:
            CC2+=i.username
            CC2+=','
    BB1=BB1.split(',')
    CC1=CC1.split(',')
    BB2=BB2.split(',')
    CC2=CC2.split(',')
    AT=A1+A2+A3+A4+A5+A6+A7+A8+A9+A10+A11+A12
    BT=B1+B2+B3+B4+B5+B6+B7+B8+B9+B10+B11+B12
    params = {
        "UserID":request.user,
        "A1":A1,"A2":A2,"A3":A3,"A4":A4,"A5":A5,"A6":A6,"A7":A7,"A8":A8,"A9":A9,"A10":A10,"A11":A11,"A12":A12,
        "B1":B1,"B2":B2,"B3":B3,"B4":B4,"B5":B5,"B6":B6,"B7":B7,"B8":B8,"B9":B9,"B10":B10,"B11":B11,"B12":B12,
        "NA1":NA1,"NA2":NA2,"NA3":NA3,"NA4":NA4,"NA5":NA5,"NA6":NA6,"NA7":NA7,"NA8":NA8,"NA9":NA9,"NA10":NA10,"NA11":NA11,"NA12":NA12,
        "NB1":NB1,"NB2":NB2,"NB3":NB3,"NB4":NB4,"NB5":NB5,"NB6":NB6,"NB7":NB7,"NB8":NB8,"NB9":NB9,"NB10":NB10,"NB11":NB11,"NB12":NB12,
        "BB1":BB1,
        "BB2":BB2,
        "CC1":CC1,
        "CC2":CC2,
        "AT":AT,
        "BT":BT,
        }
    return render(request, "blog/ADC22DL.html",context=params)


@login_required
def ADC22ESTA(request):

    colaborador2='hyuma.nozaki@cosmotec.com.br'
    resultb = ''
    resultb2 = ''
    resultb3 = ''
    resultb4 = ''
    resultb5 = ''
    resultb6 = 0
    object_list = User.objects.all()
    for l in object_list:
        resultb = l.username
        resultb2 += resultb
        resultb2 += ' '
        resultb3 = str(l.id)
        resultb4 += resultb3
        resultb4 += ' '
    resultb2=resultb2.split()
    resultb4=resultb4.split()
    m=int(len(resultb2))
    for n in range(m):
        if (str(resultb2[n])==colaborador2):
            resultb6=int(resultb4[n])

    Elist=[]  
    Elist+=RHDT.objects.values_list('ADC22E1C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E2C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E3C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E4C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E5C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E6C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E7C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E8C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E9C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E10C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E11C').get(user=resultb6)
    Elist+=RHDT.objects.values_list('ADC22E12C').get(user=resultb6)

    object_list = User.objects.all()
    
    G1C=0.0
    G2C=0.0
    G3C=0.0
    G4C=0.0
    G5C=0.0
    G6C=0.0
    G7C=0.0

    G1A=0.0
    G2A=0.0
    G3A=0.0
    G4A=0.0
    G5A=0.0
    G6A=0.0
    G7A=0.0

    G1CN=0
    G2CN=0
    G3CN=0
    G4CN=0
    G5CN=0
    G6CN=0
    G7CN=0

    G1AN=0
    G2AN=0
    G3AN=0
    G4AN=0
    G5AN=0
    G6AN=0
    G7AN=0

    G1F=0
    G2F=0
    G3F=0
    G4F=0
    G5F=0
    G6F=0
    G7F=0

    G1FC=0
    G2FC=0
    G3FC=0
    G4FC=0
    G5FC=0
    G6FC=0
    G7FC=0

    G1FA=0
    G2FA=0
    G3FA=0
    G4FA=0
    G5FA=0
    G6FA=0
    G7FA=0

    D1G1=0
    D1G1C=0
    D1G1A=0
    D1G2=0
    D1G2C=0
    D1G2A=0
    D1G3=0
    D1G3C=0
    D1G3A=0
    D1G4=0
    D1G4C=0
    D1G4A=0
    D1G5=0
    D1G5C=0
    D1G5A=0
    D1G6=0
    D1G6C=0
    D1G6A=0
    D1G7=0
    D1G7C=0
    D1G7A=0

    D2N=0
    D2G1C=0
    D2G1A=0
    D2G2C=0
    D2G2A=0
    D2G3C=0
    D2G3A=0
    D2G4C=0
    D2G4A=0
    D2G5C=0
    D2G5A=0
    D2G6C=0
    D2G6A=0
    D2G7C=0
    D2G7A=0

    D3N=0
    D3G1C=0
    D3G1A=0
    D3G2C=0
    D3G2A=0
    D3G3C=0
    D3G3A=0
    D3G4C=0
    D3G4A=0
    D3G5C=0
    D3G5A=0
    D3G6C=0
    D3G6A=0
    D3G7C=0
    D3G7A=0

    D4N=0
    D4G1C=0
    D4G1A=0
    D4G2C=0
    D4G2A=0
    D4G3C=0
    D4G3A=0
    D4G4C=0
    D4G4A=0
    D4G5C=0
    D4G5A=0
    D4G6C=0
    D4G6A=0
    D4G7C=0
    D4G7A=0

    D5N=0
    D5G1C=0
    D5G1A=0
    D5G2C=0
    D5G2A=0
    D5G3C=0
    D5G3A=0
    D5G4C=0
    D5G4A=0
    D5G5C=0
    D5G5A=0
    D5G6C=0
    D5G6A=0
    D5G7C=0
    D5G7A=0

    D6N=0
    D6G1C=0
    D6G1A=0
    D6G2C=0
    D6G2A=0
    D6G3C=0
    D6G3A=0
    D6G4C=0
    D6G4A=0
    D6G5C=0
    D6G5A=0
    D6G6C=0
    D6G6A=0
    D6G7C=0
    D6G7A=0

    D7N=0
    D7G1C=0
    D7G1A=0
    D7G2C=0
    D7G2A=0
    D7G3C=0
    D7G3A=0
    D7G4C=0
    D7G4A=0
    D7G5C=0
    D7G5A=0
    D7G6C=0
    D7G6A=0
    D7G7C=0
    D7G7A=0

    D8N=0
    D8G1C=0
    D8G1A=0
    D8G2C=0
    D8G2A=0
    D8G3C=0
    D8G3A=0
    D8G4C=0
    D8G4A=0
    D8G5C=0
    D8G5A=0
    D8G6C=0
    D8G6A=0
    D8G7C=0
    D8G7A=0

    D9N=0
    D9G1C=0
    D9G1A=0
    D9G2C=0
    D9G2A=0
    D9G3C=0
    D9G3A=0
    D9G4C=0
    D9G4A=0
    D9G5C=0
    D9G5A=0
    D9G6C=0
    D9G6A=0
    D9G7C=0
    D9G7A=0

    D10N=0
    D10G1C=0
    D10G1A=0
    D10G2C=0
    D10G2A=0
    D10G3C=0
    D10G3A=0
    D10G4C=0
    D10G4A=0
    D10G5C=0
    D10G5A=0
    D10G6C=0
    D10G6A=0
    D10G7C=0
    D10G7A=0

    D11N=0
    D11G1C=0
    D11G1A=0
    D11G2C=0
    D11G2A=0
    D11G3C=0
    D11G3A=0
    D11G4C=0
    D11G4A=0
    D11G5C=0
    D11G5A=0
    D11G6C=0
    D11G6A=0
    D11G7C=0
    D11G7A=0

    D12N=0
    D12G1C=0
    D12G1A=0
    D12G2C=0
    D12G2A=0
    D12G3C=0
    D12G3A=0
    D12G4C=0
    D12G4A=0
    D12G5C=0
    D12G5A=0
    D12G6C=0
    D12G6A=0
    D12G7C=0
    D12G7A=0

    D13N=0
    D13G1C=0
    D13G1A=0
    D13G2C=0
    D13G2A=0
    D13G3C=0
    D13G3A=0
    D13G4C=0
    D13G4A=0
    D13G5C=0
    D13G5A=0
    D13G6C=0
    D13G6A=0
    D13G7C=0
    D13G7A=0

    D14N=0
    D14G1C=0
    D14G1A=0
    D14G2C=0
    D14G2A=0
    D14G3C=0
    D14G3A=0
    D14G4C=0
    D14G4A=0
    D14G5C=0
    D14G5A=0
    D14G6C=0
    D14G6A=0
    D14G7C=0
    D14G7A=0

    D15N=0
    D15G1C=0
    D15G1A=0
    D15G2C=0
    D15G2A=0
    D15G3C=0
    D15G3A=0
    D15G4C=0
    D15G4A=0
    D15G5C=0
    D15G5A=0
    D15G6C=0
    D15G6A=0
    D15G7C=0
    D15G7A=0

    D16N=0
    D16G1C=0
    D16G1A=0
    D16G2C=0
    D16G2A=0
    D16G3C=0
    D16G3A=0
    D16G4C=0
    D16G4A=0
    D16G5C=0
    D16G5A=0
    D16G6C=0
    D16G6A=0
    D16G7C=0
    D16G7A=0

    D1=''
    D2=''
    D3=''
    D4=''
    D5=''
    D6=''
    D7=''
    D8=''
    D9=''
    D10=''
    D11=''
    D12=''
    D13=''
    D14=''
    D15=''
    D16=''

    E1C=0.0
    E1N=0
    E2C=0.0
    E2N=0
    E3C=0.0
    E3N=0
    E4C=0.0
    E4N=0
    E5C=0.0
    E5N=0
    E6C=0.0
    E6N=0
    E7C=0.0
    E7N=0
    E8C=0.0
    E8N=0
    E9C=0.0
    E9N=0
    E10C=0.0
    E10N=0
    E11C=0.0
    E11N=0
    E12C=0.0
    E12N=0
    E13C=0.0
    E13N=0
    E14C=0.0
    E14N=0
    E15C=0.0
    E15N=0
    E16C=0.0
    E16N=0
    E1A=0.0
    E2A=0.0
    E3A=0.0
    E4A=0.0
    E5A=0.0
    E6A=0.0
    E7A=0.0
    E8A=0.0
    E9A=0.0
    E10A=0.0
    E11A=0.0
    E12A=0.0
    E13A=0.0
    E14A=0.0
    E15A=0.0
    E16A=0.0

    for i in object_list:
        checkkk = []
        checkkk += ADC22.objects.values_list('ADC22A').get(user=i.id)
        if int(checkkk[0] == 1):
            try:
                CE1=[]
                #CE1+=ADC22.objects.values_list('ADC22E1').get(user=i.id)
                CE1.append(i.first_name)
                if Elist[0]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E1C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E1A+=float(AA[1])
                    E1N+=1
                if Elist[1]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E2C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E2A+=float(AA[1])
                    E2N+=1
                if Elist[2]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E3C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E3A+=float(AA[1])
                    E3N+=1
                if Elist[3]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E4C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E4A+=float(AA[1])
                    E4N+=1
                if Elist[4]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E5C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E5A+=float(AA[1])
                    E5N+=1
                if Elist[5]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E6C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E6A+=float(AA[1])
                    E6N+=1
                if Elist[6]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E7C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E7A+=float(AA[1])
                    E7N+=1
                if Elist[7]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E8C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E8A+=float(AA[1])
                    E8N+=1
                if Elist[8]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E9C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E9A+=float(AA[1])
                    E9N+=1
                if Elist[9]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E10C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E10A+=float(AA[1])
                    E10N+=1
                if Elist[10]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E11C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E11A+=float(AA[1])
                    E11N+=1
                if Elist[11]==CE1[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
                    E12C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
                    E12A+=float(AA[1])
                    E12N+=1
                CE2=[]
#            CE2+=ADC22.objects.values_list('ADC22E2').get(user=i.id)
                CE2.append(i.last_name)
                if Elist[0]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E1C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E1A+=float(AA[1])
                    E1N+=1
                if Elist[1]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E2C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E2A+=float(AA[1])
                    E2N+=1
                if Elist[2]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E3C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E3A+=float(AA[1])
                    E3N+=1
                if Elist[3]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E4C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E4A+=float(AA[1])
                    E4N+=1
                if Elist[4]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E5C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E5A+=float(AA[1])
                    E5N+=1
                if Elist[5]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E6C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E6A+=float(AA[1])
                    E6N+=1
                if Elist[6]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E7C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E7A+=float(AA[1])
                    E7N+=1
                if Elist[7]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E8C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E8A+=float(AA[1])
                    E8N+=1
                if Elist[8]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E9C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E9A+=float(AA[1])
                    E9N+=1
                if Elist[9]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E10C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E10A+=float(AA[1])
                    E10N+=1
                if Elist[10]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E11C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E11A+=float(AA[1])
                    E11N+=1
                if Elist[11]==CE2[0]:
                    AA=[]
                    AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
                    E12C+=float(AA[0])
                    AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
                    E12A+=float(AA[1])
                    E12N+=1
            except :
                AA=[]

    for i in object_list:
        checkkk = []
        checkkk += ADC22.objects.values_list('ADC22A').get(user=i.id)       
        try:            
            if int(checkkk[0]) == 1:
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                G1C+=float(AA[0])
                G1CN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                G1A+=float(AA[0])
                G1AN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                G2C+=float(AA[0])
                G2CN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                G2A+=float(AA[0])
                G2AN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                G3C+=float(AA[0])
                G3CN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                G3A+=float(AA[0])
                G3AN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                G4C+=float(AA[0])
                G4CN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                G4A+=float(AA[0])
                G4AN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                G5C+=float(AA[0])
                G5CN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                G5A+=float(AA[0])
                G5AN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                G6C+=float(AA[0])
                G6CN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                G6A+=float(AA[0])
                G6AN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                G7C+=float(AA[0])
                G7CN+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                G7A+=float(AA[0])
                G7AN+=1
            if 'resultado' in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
#            if PDI22.objects.values_list('PDI22G1C').get(user=i.id)==RHDT.objects.values_list('ADC22G1').get(user=resultb6):
                G1F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                G1FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                G1FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G1C').get(user=i.id)==RHDT.objects.values_list('ADC22G2').get(user=resultb6):
            if 'cliente' in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
                G2F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                G2FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                G2FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G1C').get(user=i.id)==RHDT.objects.values_list('ADC22G3').get(user=resultb6):
            if 'busca' in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
                G3F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                G3FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                G3FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G1C').get(user=i.id)==RHDT.objects.values_list('ADC22G4').get(user=resultb6):
            if 'liderança' in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
                G4F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                G4FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                G4FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G1C').get(user=i.id)==RHDT.objects.values_list('ADC22G5').get(user=resultb6):
            if 'senso' in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
                G5F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                G5FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                G5FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G1C').get(user=i.id)==RHDT.objects.values_list('ADC22G6').get(user=resultb6):
            if 'comunicação' in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
                G6F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                G6FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                G6FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G1C').get(user=i.id)==RHDT.objects.values_list('ADC22G7').get(user=resultb6):
            if 'relacionamento' in str(PDI22.objects.values_list('PDI22G1C').get(user=i.id)).lower():
                G7F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                G7FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                G7FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G2C').get(user=i.id)==RHDT.objects.values_list('ADC22G1').get(user=resultb6):
            if 'resultado' in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                G1F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                G1FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                G1FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G2C').get(user=i.id)==RHDT.objects.values_list('ADC22G2').get(user=resultb6):
            if 'cliente' in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                G2F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                G2FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                G2FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G2C').get(user=i.id)==RHDT.objects.values_list('ADC22G3').get(user=resultb6):
            if 'busca' in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                G3F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                G3FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                G3FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G2C').get(user=i.id)==RHDT.objects.values_list('ADC22G4').get(user=resultb6):
            if 'liderança' in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                G4F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                G4FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                G4FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G2C').get(user=i.id)==RHDT.objects.values_list('ADC22G5').get(user=resultb6):
            if 'senso' in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                G5F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                G5FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                G5FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G2C').get(user=i.id)==RHDT.objects.values_list('ADC22G6').get(user=resultb6):
            if 'comunicação' in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                G6FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                G6FA+=float(AA[1])
                G6F+=1
#            if PDI22.objects.values_list('PDI22G2C').get(user=i.id)==RHDT.objects.values_list('ADC22G7').get(user=resultb6):
            if 'relacionamento' in str(PDI22.objects.values_list('PDI22G2C').get(user=i.id)).lower():
                G7F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                G7FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                G7FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G3C').get(user=i.id)==RHDT.objects.values_list('ADC22G1').get(user=resultb6):
            if 'resultado' in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                G1F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                G1FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                G1FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G3C').get(user=i.id)==RHDT.objects.values_list('ADC22G2').get(user=resultb6):
            if 'cliente' in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                G2F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                G2FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                G2FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G3C').get(user=i.id)==RHDT.objects.values_list('ADC22G3').get(user=resultb6):
            if 'busca' in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                G3F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                G3FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                G3FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G3C').get(user=i.id)==RHDT.objects.values_list('ADC22G4').get(user=resultb6):
            if 'liderança' in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                G4F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                G4FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                G4FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G3C').get(user=i.id)==RHDT.objects.values_list('ADC22G5').get(user=resultb6):
            if 'senso' in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                G5F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                G5FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                G5FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G3C').get(user=i.id)==RHDT.objects.values_list('ADC22G6').get(user=resultb6):
            if 'comunicação' in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                G6F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                G6FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                G6FA+=float(AA[1])
#            if PDI22.objects.values_list('PDI22G3C').get(user=i.id)==RHDT.objects.values_list('ADC22G7').get(user=resultb6):
            if 'relacionamento' in str(PDI22.objects.values_list('PDI22G3C').get(user=i.id)).lower():
                G7F+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                G7FC+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                G7FA+=float(AA[1])
            test=[]
            test+=MBO22.objects.values_list('DEPT').get(user=i.id)
            if test[0]=='MARKETING':
                D1=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D1G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D1G1A+=float(AA[1])
                D1G1+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D1G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D1G2A+=float(AA[1])
                D1G2+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D1G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D1G3A+=float(AA[1])
                D1G3+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D1G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D1G4A+=float(AA[1])
                D1G4+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D1G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D1G5A+=float(AA[1])
                D1G5+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D1G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D1G6A+=float(AA[1])
                D1G6+=1
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D1G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D1G7A+=float(AA[1])
                D1G7+=1
            if test[0]=='GERENCIA DE PRODUTOS E FORNECEDORES':
                D2N+=1
                D2=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D2G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D2G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D2G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D2G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D2G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D2G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D2G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D2G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D2G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D2G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D2G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D2G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D2G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D2G7A+=float(AA[1])
            if test[0]=='COMERCIAL':
                D3N+=1
                D3=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D3G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D3G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D3G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D3G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D3G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D3G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D3G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D3G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D3G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D3G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D3G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D3G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D3G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D3G7A+=float(AA[1])
            if test[0]=='OPERACIONAL':
                D4N+=1
                D4=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D4G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D4G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D4G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D4G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D4G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D4G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D4G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D4G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D4G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D4G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D4G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D4G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D4G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D4G7A+=float(AA[1])
            if test[0]=='ADMINISTRATIVO':
                D5N+=1
                D5=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D5G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D5G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D5G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D5G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D5G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D5G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D5G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D5G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D5G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D5G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D5G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D5G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D5G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D5G7A+=float(AA[1])

            if test[0]=='COMERCIAL EXPRESS':
                D6N+=1
                D6=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D6G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D6G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D6G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D6G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D6G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D6G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D6G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D6G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D6G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D6G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D6G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D6G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D6G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D6G7A+=float(AA[1])

            if test[0]=='CUSTOMER SERVICE':
                D7N+=1
                D7=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D7G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D7G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D7G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D7G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D7G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D7G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D7G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D7G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D7G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D7G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D7G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D7G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D7G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D7G7A+=float(AA[1])

            if test[0]=='GARANTIA DA QUALIDADE':
                D8N+=1
                D8=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D8G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D8G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D8G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D8G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D8G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D8G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D8G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D8G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D8G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D8G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D8G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D8G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D8G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D8G7A+=float(AA[1])

            if test[0]=='LOGISTICA':
                D9N+=1
                D9=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D9G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D9G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D9G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D9G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D9G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D9G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D9G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D9G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D9G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D9G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D9G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D9G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D9G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D9G7A+=float(AA[1])

            if test[0]=='PESQUISA E DESENVOLVIMENTO':
                D10N+=1
                D10=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D10G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D10G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D10G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D10G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D10G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D10G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D10G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D10G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D10G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D10G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D10G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D10G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D10G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D10G7A+=float(AA[1])

            if test[0]=='PLANEJAMENTO DE MATERIAIS':
                D11N+=1
                D11=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D11G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D11G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D11G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D11G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D11G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D11G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D11G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D11G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D11G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D11G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D11G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D11G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D11G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D11G7A+=float(AA[1])

            if test[0]=='PLANEJAMENTO FINANCEIRO - FP&A':
                D12N+=1
                D12=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D12G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D12G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D12G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D12G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D12G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D12G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D12G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D12G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D12G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D12G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D12G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D12G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D12G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D12G7A+=float(AA[1])

            if test[0]=='RECURSOS HUMANOS':
                D13N+=1
                D13=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D13G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D13G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D13G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D13G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D13G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D13G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D13G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D13G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D13G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D13G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D13G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D13G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D13G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D13G7A+=float(AA[1])


            if test[0]=='TECNOLOGIA DA INFORMACAO':
                D14N+=1
                D14=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D14G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D14G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D14G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D14G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D14G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D14G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D14G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D14G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D14G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D14G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D14G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D14G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D14G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D14G7A+=float(AA[1])

            if test[0]=='UNIDADE FRAGRANCIAS':
                D15N+=1
                D15=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D15G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D15G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D15G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D15G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D15G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D15G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D15G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D15G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D15G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D15G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D15G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D15G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D15G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D15G7A+=float(AA[1])


            if test[0]=='Financeiro':
                D16N+=1
                D16=test[0]
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
                D16G1C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
                D16G1A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
                D16G2C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
                D16G2A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
                D16G3C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
                D16G3A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
                D16G4C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
                D16G4A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
                D16G5C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
                D16G5A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
                D16G6C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
                D16G6A+=float(AA[1])
                AA=[]
                AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
                D16G7C+=float(AA[0])
                AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
                D16G7A+=float(AA[1])


        except:
            AA=1

    if E1N!=0:
        E1C=round(E1C/E1N,1)
        E1A=round(E1A/E1N,1)
    if E2N!=0:
        E2C=round(E2C/E2N,1)
        E2A=round(E2A/E2N,1)
    if E3N!=0:
        E3C=round(E3C/E3N,1)
        E3A=round(E3A/E3N,1)
    if E4N!=0:
        E4C=round(E4C/E4N,1)
        E4A=round(E4A/E4N,1)
    if E5N!=0:
        E5C=round(E5C/E5N,1)
        E5A=round(E5A/E5N,1)
    if E6N!=0:
        E6C=round(E6C/E6N,1)
        E6A=round(E6A/E6N,1)
    if E7N!=0:
        E7C=round(E7C/E7N,1)
        E7A=round(E7A/E7N,1)
    if E8N!=0:
        E8C=round(E8C/E8N,1)
        E8A=round(E8A/E8N,1)
    if E9N!=0:
        E9C=round(E9C/E9N,1)
        E9A=round(E9A/E9N,1)
    if E10N!=0:
        E10C=round(E10C/E10N,1)
        E10A=round(E10A/E10N,1)
    if E11N!=0:
        E11C=round(E11C/E11N,1)
        E11A=round(E11A/E11N,1)
    if E12N!=0:
        E12C=round(E12C/E12N,1)
        E12A=round(E12A/E12N,1)


    G1C=round(G1C/G1CN,1)
    G1A=round(G1A/G1AN,1)
    G2C=round(G2C/G2CN,1)
    G2A=round(G2A/G2AN,1)
    G3C=round(G3C/G3CN,1)
    G3A=round(G3A/G3AN,1)
    G4C=round(G4C/G4CN,1)
    G4A=round(G4A/G4AN,1)
    G5C=round(G5C/G5CN,1)
    G5A=round(G5A/G5AN,1)
    G6C=round(G6C/G6CN,1)
    G6A=round(G6A/G6AN,1)
    G7C=round(G7C/G7CN,1)
    G7A=round(G7A/G7AN,1)

    if G1F!=0:
        G1FC=round(G1FC/G1F,1)
        G1FA=round(G1FA/G1F,1)
    if G2F!=0:
        G2FC=round(G2FC/G2F,1)
        G2FA=round(G2FA/G2F,1)
    if G3F!=0:
        G3FC=round(G3FC/G3F,1)
        G3FA=round(G3FA/G3F,1)
    if G4F!=0:
        G4FC=round(G4FC/G4F,1)
        G4FA=round(G4FA/G4F,1)
    if G5F!=0:
        G5FC=round(G5FC/G5F,1)
        G5FA=round(G5FA/G5F,1)
    if G6F!=0:
        G6FC=round(G6FC/G6F,1)
        G6FA=round(G6FA/G6F,1)
    if G7F!=0:
        G7FC=round(G7FC/G7F,1)
        G7FA=round(G7FA/G7F,1)

    if D1G1==0:
        D1G1=0
        D1G1C=0
        D1G1A=0
    else:
        D1G1C=round(D1G1C/D1G1,1)
        D1G1A=round(D1G1A/D1G1,1)
    if D1G2==0:
        D1G2=0
        D1G2C=0
        D1G2A=0
    else:
        D1G2C=round(D1G2C/D1G2,1)
        D1G2A=round(D1G2A/D1G2,1)
    if D1G3==0:
        D1G3=0
        D1G3C=0
        D1G3A=0
    else:
        D1G3C=round(D1G3C/D1G3,1)
        D1G3A=round(D1G3A/D1G3,1)
    if D1G4==0:
        D1G4=0
        D1G4C=0
        D1G4A=0
    else:
        D1G4C=round(D1G4C/D1G4,1)
        D1G4A=round(D1G4A/D1G4,1)
    if D1G5==0:
        D1G5=0
        D1G5C=0
        D1G5A=0
    else:
        D1G5C=round(D1G5C/D1G5,1)
        D1G5A=round(D1G5A/D1G5,1)
    if D1G6==0:
        D1G6=0
        D1G6C=0
        D1G6A=0
    else:
        D1G6C=round(D1G6C/D1G6,1)
        D1G6A=round(D1G6A/D1G6,1)
    if D1G7==0:
        D1G7=0
        D1G7C=0
        D1G7A=0
    else:
        D1G7C=round(D1G7C/D1G7,1)
        D1G7A=round(D1G7A/D1G7,1)

    if D2N!=0:
        D2G1C=round(D2G1C/D2N,1)
        D2G1A=round(D2G1A/D2N,1)
        D2G2C=round(D2G2C/D2N,1)
        D2G2A=round(D2G2A/D2N,1)
        D2G3C=round(D2G3C/D2N,1)
        D2G3A=round(D2G3A/D2N,1)
        D2G4C=round(D2G4C/D2N,1)
        D2G4A=round(D2G4A/D2N,1)
        D2G5C=round(D2G5C/D2N,1)
        D2G5A=round(D2G5A/D2N,1)
        D2G6C=round(D2G6C/D2N,1)
        D2G6A=round(D2G6A/D2N,1)
        D2G7C=round(D2G7C/D2N,1)
        D2G7A=round(D2G7A/D2N,1)

    if D3N!=0:
        D3G1C=round(D3G1C/D3N,1)
        D3G1A=round(D3G1A/D3N,1)
        D3G2C=round(D3G2C/D3N,1)
        D3G2A=round(D3G2A/D3N,1)
        D3G3C=round(D3G3C/D3N,1)
        D3G3A=round(D3G3A/D3N,1)
        D3G4C=round(D3G4C/D3N,1)
        D3G4A=round(D3G4A/D3N,1)
        D3G5C=round(D3G5C/D3N,1)
        D3G5A=round(D3G5A/D3N,1)
        D3G6C=round(D3G6C/D3N,1)
        D3G6A=round(D3G6A/D3N,1)
        D3G7C=round(D3G7C/D3N,1)
        D3G7A=round(D3G7A/D3N,1)

    if D4N!=0:
        D4G1C=round(D4G1C/D4N,1)
        D4G1A=round(D4G1A/D4N,1)
        D4G2C=round(D4G2C/D4N,1)
        D4G2A=round(D4G2A/D4N,1)
        D4G3C=round(D4G3C/D4N,1)
        D4G3A=round(D4G3A/D4N,1)
        D4G4C=round(D4G4C/D4N,1)
        D4G4A=round(D4G4A/D4N,1)
        D4G5C=round(D4G5C/D4N,1)
        D4G5A=round(D4G5A/D4N,1)
        D4G6C=round(D4G6C/D4N,1)
        D4G6A=round(D4G6A/D4N,1)
        D4G7C=round(D4G7C/D4N,1)
        D4G7A=round(D4G7A/D4N,1)

    if D5N!=0:
        D5G1C=round(D5G1C/D5N,1)
        D5G1A=round(D5G1A/D5N,1)
        D5G2C=round(D5G2C/D5N,1)
        D5G2A=round(D5G2A/D5N,1)
        D5G3C=round(D5G3C/D5N,1)
        D5G3A=round(D5G3A/D5N,1)
        D5G4C=round(D5G4C/D5N,1)
        D5G4A=round(D5G4A/D5N,1)
        D5G5C=round(D5G5C/D5N,1)
        D5G5A=round(D5G5A/D5N,1)
        D5G6C=round(D5G6C/D5N,1)
        D5G6A=round(D5G6A/D5N,1)
        D5G7C=round(D5G7C/D5N,1)
        D5G7A=round(D5G7A/D5N,1)


    if D6N!=0:
        D6G1C=round(D6G1C/D6N,1)
        D6G1A=round(D6G1A/D6N,1)
        D6G2C=round(D6G2C/D6N,1)
        D6G2A=round(D6G2A/D6N,1)
        D6G3C=round(D6G3C/D6N,1)
        D6G3A=round(D6G3A/D6N,1)
        D6G4C=round(D6G4C/D6N,1)
        D6G4A=round(D6G4A/D6N,1)
        D6G5C=round(D6G5C/D6N,1)
        D6G5A=round(D6G5A/D6N,1)
        D6G6C=round(D6G6C/D6N,1)
        D6G6A=round(D6G6A/D6N,1)
        D6G7C=round(D6G7C/D6N,1)
        D6G7A=round(D6G7A/D6N,1)

    if D7N!=0:
        D7G1C=round(D7G1C/D7N,1)
        D7G1A=round(D7G1A/D7N,1)
        D7G2C=round(D7G2C/D7N,1)
        D7G2A=round(D7G2A/D7N,1)
        D7G3C=round(D7G3C/D7N,1)
        D7G3A=round(D7G3A/D7N,1)
        D7G4C=round(D7G4C/D7N,1)
        D7G4A=round(D7G4A/D7N,1)
        D7G5C=round(D7G5C/D7N,1)
        D7G5A=round(D7G5A/D7N,1)
        D7G6C=round(D7G6C/D7N,1)
        D7G6A=round(D7G6A/D7N,1)
        D7G7C=round(D7G7C/D7N,1)
        D7G7A=round(D7G7A/D7N,1)

    if D8N!=0:
        D8G1C=round(D8G1C/D8N,1)
        D8G1A=round(D8G1A/D8N,1)
        D8G2C=round(D8G2C/D8N,1)
        D8G2A=round(D8G2A/D8N,1)
        D8G3C=round(D8G3C/D8N,1)
        D8G3A=round(D8G3A/D8N,1)
        D8G4C=round(D8G4C/D8N,1)
        D8G4A=round(D8G4A/D8N,1)
        D8G5C=round(D8G5C/D8N,1)
        D8G5A=round(D8G5A/D8N,1)
        D8G6C=round(D8G6C/D8N,1)
        D8G6A=round(D8G6A/D8N,1)
        D8G7C=round(D8G7C/D8N,1)
        D8G7A=round(D8G7A/D8N,1)

    if D9N!=0:
        D9G1C=round(D9G1C/D9N,1)
        D9G1A=round(D9G1A/D9N,1)
        D9G2C=round(D9G2C/D9N,1)
        D9G2A=round(D9G2A/D9N,1)
        D9G3C=round(D9G3C/D9N,1)
        D9G3A=round(D9G3A/D9N,1)
        D9G4C=round(D9G4C/D9N,1)
        D9G4A=round(D9G4A/D9N,1)
        D9G5C=round(D9G5C/D9N,1)
        D9G5A=round(D9G5A/D9N,1)
        D9G6C=round(D9G6C/D9N,1)
        D9G6A=round(D9G6A/D9N,1)
        D9G7C=round(D9G7C/D9N,1)
        D9G7A=round(D9G7A/D9N,1)

    if D10N!=0:
        D10G1C=round(D10G1C/D10N,1)
        D10G1A=round(D10G1A/D10N,1)
        D10G2C=round(D10G2C/D10N,1)
        D10G2A=round(D10G2A/D10N,1)
        D10G3C=round(D10G3C/D10N,1)
        D10G3A=round(D10G3A/D10N,1)
        D10G4C=round(D10G4C/D10N,1)
        D10G4A=round(D10G4A/D10N,1)
        D10G5C=round(D10G5C/D10N,1)
        D10G5A=round(D10G5A/D10N,1)
        D10G6C=round(D10G6C/D10N,1)
        D10G6A=round(D10G6A/D10N,1)
        D10G7C=round(D10G7C/D10N,1)
        D10G7A=round(D10G7A/D10N,1)

    if D11N!=0:
        D11G1C=round(D11G1C/D11N,1)
        D11G1A=round(D11G1A/D11N,1)
        D11G2C=round(D11G2C/D11N,1)
        D11G2A=round(D11G2A/D11N,1)
        D11G3C=round(D11G3C/D11N,1)
        D11G3A=round(D11G3A/D11N,1)
        D11G4C=round(D11G4C/D11N,1)
        D11G4A=round(D11G4A/D11N,1)
        D11G5C=round(D11G5C/D11N,1)
        D11G5A=round(D11G5A/D11N,1)
        D11G6C=round(D11G6C/D11N,1)
        D11G6A=round(D11G6A/D11N,1)
        D11G7C=round(D11G7C/D11N,1)
        D11G7A=round(D11G7A/D11N,1)

    if D12N!=0:
        D12G1C=round(D12G1C/D12N,1)
        D12G1A=round(D12G1A/D12N,1)
        D12G2C=round(D12G2C/D12N,1)
        D12G2A=round(D12G2A/D12N,1)
        D12G3C=round(D12G3C/D12N,1)
        D12G3A=round(D12G3A/D12N,1)
        D12G4C=round(D12G4C/D12N,1)
        D12G4A=round(D12G4A/D12N,1)
        D12G5C=round(D12G5C/D12N,1)
        D12G5A=round(D12G5A/D12N,1)
        D12G6C=round(D12G6C/D12N,1)
        D12G6A=round(D12G6A/D12N,1)
        D12G7C=round(D12G7C/D12N,1)
        D12G7A=round(D12G7A/D12N,1)

    if D13N!=0:
        D13G1C=round(D13G1C/D13N,1)
        D13G1A=round(D13G1A/D13N,1)
        D13G2C=round(D13G2C/D13N,1)
        D13G2A=round(D13G2A/D13N,1)
        D13G3C=round(D13G3C/D13N,1)
        D13G3A=round(D13G3A/D13N,1)
        D13G4C=round(D13G4C/D13N,1)
        D13G4A=round(D13G4A/D13N,1)
        D13G5C=round(D13G5C/D13N,1)
        D13G5A=round(D13G5A/D13N,1)
        D13G6C=round(D13G6C/D13N,1)
        D13G6A=round(D13G6A/D13N,1)
        D13G7C=round(D13G7C/D13N,1)
        D13G7A=round(D13G7A/D13N,1)

    if D14N!=0:
        D14G1C=round(D14G1C/D14N,1)
        D14G1A=round(D14G1A/D14N,1)
        D14G2C=round(D14G2C/D14N,1)
        D14G2A=round(D14G2A/D14N,1)
        D14G3C=round(D14G3C/D14N,1)
        D14G3A=round(D14G3A/D14N,1)
        D14G4C=round(D14G4C/D14N,1)
        D14G4A=round(D14G4A/D14N,1)
        D14G5C=round(D14G5C/D14N,1)
        D14G5A=round(D14G5A/D14N,1)
        D14G6C=round(D14G6C/D14N,1)
        D14G6A=round(D14G6A/D14N,1)
        D14G7C=round(D14G7C/D14N,1)
        D14G7A=round(D14G7A/D14N,1)

    if D15N!=0:
        D15G1C=round(D15G1C/D15N,1)
        D15G1A=round(D15G1A/D15N,1)
        D15G2C=round(D15G2C/D15N,1)
        D15G2A=round(D15G2A/D15N,1)
        D15G3C=round(D15G3C/D15N,1)
        D15G3A=round(D15G3A/D15N,1)
        D15G4C=round(D15G4C/D15N,1)
        D15G4A=round(D15G4A/D15N,1)
        D15G5C=round(D15G5C/D15N,1)
        D15G5A=round(D15G5A/D15N,1)
        D15G6C=round(D15G6C/D15N,1)
        D15G6A=round(D15G6A/D15N,1)
        D15G7C=round(D15G7C/D15N,1)
        D15G7A=round(D15G7A/D15N,1)

    if D16N!=0:
        D16G1C=round(D16G1C/D16N,1)
        D16G1A=round(D16G1A/D16N,1)
        D16G2C=round(D16G2C/D16N,1)
        D16G2A=round(D16G2A/D16N,1)
        D16G3C=round(D16G3C/D16N,1)
        D16G3A=round(D16G3A/D16N,1)
        D16G4C=round(D16G4C/D16N,1)
        D16G4A=round(D16G4A/D16N,1)
        D16G5C=round(D16G5C/D16N,1)
        D16G5A=round(D16G5A/D16N,1)
        D16G6C=round(D16G6C/D16N,1)
        D16G6A=round(D16G6A/D16N,1)
        D16G7C=round(D16G7C/D16N,1)
        D16G7A=round(D16G7A/D16N,1)

    GG=[]
    GG+=RHDT.objects.values_list('ADC22G1').get(user=resultb6)
    GG+=RHDT.objects.values_list('ADC22G2').get(user=resultb6)
    GG+=RHDT.objects.values_list('ADC22G3').get(user=resultb6)
    GG+=RHDT.objects.values_list('ADC22G4').get(user=resultb6)
    GG+=RHDT.objects.values_list('ADC22G5').get(user=resultb6)
    GG+=RHDT.objects.values_list('ADC22G6').get(user=resultb6)
    GG+=RHDT.objects.values_list('ADC22G7').get(user=resultb6)

    params = {
        "test":test,
        "UserID":request.user,
        "G1":GG[0],
        "G2":GG[1],
        "G3":GG[2],
        "G4":GG[3],
        "G5":GG[4],
        "G6":GG[5],
        "G7":GG[6],
        "G1C":G1C,
        "G1A":G1A,
        "G1CN":G1CN,
        "G1AN":G1AN,
        "G2C":G2C,
        "G2A":G2A,
        "G2CN":G2CN,
        "G2AN":G2AN,
        "G3C":G3C,
        "G3A":G3A,
        "G3CN":G3CN,
        "G3AN":G3AN,
        "G4C":G4C,
        "G4A":G4A,
        "G4CN":G4CN,
        "G4AN":G4AN,
        "G5C":G5C,
        "G5A":G5A,
        "G5CN":G5CN,
        "G5AN":G5AN,
        "G6C":G6C,
        "G6A":G6A,
        "G6CN":G6CN,
        "G6AN":G6AN,
        "G7C":G7C,
        "G7A":G7A,
        "G7CN":G7CN,
        "G7AN":G7AN,
        "G1FC":G1FC,
        "G2FC":G2FC,
        "G3FC":G3FC,
        "G4FC":G4FC,
        "G5FC":G5FC,
        "G6FC":G6FC,
        "G7FC":G7FC,
        "G1FA":G1FA,
        "G2FA":G2FA,
        "G3FA":G3FA,
        "G4FA":G4FA,
        "G5FA":G5FA,
        "G6FA":G6FA,
        "G7FA":G7FA,
        "G1F":G1F,
        "G2F":G2F,
        "G3F":G3F,
        "G4F":G4F,
        "G5F":G5F,
        "G6F":G6F,
        "G7F":G7F,
        "D1":D1,
        "D1N":D1G1,
        "D1G1C":D1G1C,
        "D1G1A":D1G1A,
        "D1G2C":D1G2C,
        "D1G2A":D1G2A,
        "D1G3C":D1G3C,
        "D1G3A":D1G3A,
        "D1G4C":D1G4C,
        "D1G4A":D1G4A,
        "D1G5C":D1G5C,
        "D1G5A":D1G5A,
        "D1G6C":D1G6C,
        "D1G6A":D1G6A,
        "D1G7C":D1G7C,
        "D1G7A":D1G7A,
        "D2":D2,
        "D2N":D2N,
        "D2G1C":D2G1C,
        "D2G1A":D2G1A,
        "D2G2C":D2G2C,
        "D2G2A":D2G2A,
        "D2G3C":D2G3C,
        "D2G3A":D2G3A,
        "D2G4C":D2G4C,
        "D2G4A":D2G4A,
        "D2G5C":D2G5C,
        "D2G5A":D2G5A,
        "D2G6C":D2G6C,
        "D2G6A":D2G6A,
        "D2G7C":D2G7C,
        "D2G7A":D2G7A,
        "D3":D3,
        "D3N":D3N,
        "D3G1C":D3G1C,
        "D3G1A":D3G1A,
        "D3G2C":D3G2C,
        "D3G2A":D3G2A,
        "D3G3C":D3G3C,
        "D3G3A":D3G3A,
        "D3G4C":D3G4C,
        "D3G4A":D3G4A,
        "D3G5C":D3G5C,
        "D3G5A":D3G5A,
        "D3G6C":D3G6C,
        "D3G6A":D3G6A,
        "D3G7C":D3G7C,
        "D3G7A":D3G7A,
        "D4":D4,
        "D4N":D4N,
        "D4G1C":D4G1C,
        "D4G1A":D4G1A,
        "D4G2C":D4G2C,
        "D4G2A":D4G2A,
        "D4G3C":D4G3C,
        "D4G3A":D4G3A,
        "D4G4C":D4G4C,
        "D4G4A":D4G4A,
        "D4G5C":D4G5C,
        "D4G5A":D4G5A,
        "D4G6C":D4G6C,
        "D4G6A":D4G6A,
        "D4G7C":D4G7C,
        "D4G7A":D4G7A,
        "D5":D5,
        "D5N":D5N,
        "D5G1C":D5G1C,
        "D5G1A":D5G1A,
        "D5G2C":D5G2C,
        "D5G2A":D15G2A,
        "D5G3C":D5G3C,
        "D5G3A":D5G3A,
        "D5G4C":D5G4C,
        "D5G4A":D5G4A,
        "D5G5C":D5G5C,
        "D5G5A":D5G5A,
        "D5G6C":D5G6C,
        "D5G6A":D5G6A,
        "D5G7C":D5G7C,
        "D5G7A":D5G7A,
        "D6":D6,
        "D6N":D6N,
        "D6G1C":D6G1C,
        "D6G1A":D6G1A,
        "D6G2C":D6G2C,
        "D6G2A":D6G2A,
        "D6G3C":D6G3C,
        "D6G3A":D6G3A,
        "D6G4C":D6G4C,
        "D6G4A":D6G4A,
        "D6G5C":D6G5C,
        "D6G5A":D6G5A,
        "D6G6C":D6G6C,
        "D6G6A":D6G6A,
        "D6G7C":D6G7C,
        "D6G7A":D6G7A,
        "D7":D7,
        "D7N":D7N,
        "D7G1C":D7G1C,
        "D7G1A":D7G1A,
        "D7G2C":D7G2C,
        "D7G2A":D7G2A,
        "D7G3C":D7G3C,
        "D7G3A":D7G3A,
        "D7G4C":D7G4C,
        "D7G4A":D7G4A,
        "D7G5C":D7G5C,
        "D7G5A":D7G5A,
        "D7G6C":D7G6C,
        "D7G6A":D7G6A,
        "D7G7C":D7G7C,
        "D7G7A":D7G7A,
        "D8":D8,
        "D8N":D8N,
        "D8G1C":D8G1C,
        "D8G1A":D8G1A,
        "D8G2C":D8G2C,
        "D8G2A":D8G2A,
        "D8G3C":D8G3C,
        "D8G3A":D8G3A,
        "D8G4C":D8G4C,
        "D8G4A":D8G4A,
        "D8G5C":D8G5C,
        "D8G5A":D8G5A,
        "D8G6C":D8G6C,
        "D8G6A":D8G6A,
        "D8G7C":D8G7C,
        "D8G7A":D8G7A,
        "D9":D9,
        "D9N":D9N,
        "D9G1C":D9G1C,
        "D9G1A":D9G1A,
        "D9G2C":D9G2C,
        "D9G2A":D9G2A,
        "D9G3C":D9G3C,
        "D9G3A":D9G3A,
        "D9G4C":D9G4C,
        "D9G4A":D9G4A,
        "D9G5C":D9G5C,
        "D9G5A":D9G5A,
        "D9G6C":D9G6C,
        "D9G6A":D9G6A,
        "D9G7C":D9G7C,
        "D9G7A":D9G7A,
        "D10":D10,
        "D10N":D10N,
        "D10G1C":D10G1C,
        "D10G1A":D10G1A,
        "D10G2C":D10G2C,
        "D10G2A":D10G2A,
        "D10G3C":D10G3C,
        "D10G3A":D10G3A,
        "D10G4C":D10G4C,
        "D10G4A":D10G4A,
        "D10G5C":D10G5C,
        "D10G5A":D10G5A,
        "D10G6C":D10G6C,
        "D10G6A":D10G6A,
        "D10G7C":D10G7C,
        "D10G7A":D10G7A,

        "D11":D11,
        "D11N":D11N,
        "D11G1C":D11G1C,
        "D11G1A":D11G1A,
        "D11G2C":D11G2C,
        "D11G2A":D11G2A,
        "D11G3C":D11G3C,
        "D11G3A":D11G3A,
        "D11G4C":D11G4C,
        "D11G4A":D11G4A,
        "D11G5C":D11G5C,
        "D11G5A":D11G5A,
        "D11G6C":D11G6C,
        "D11G6A":D11G6A,
        "D11G7C":D11G7C,
        "D11G7A":D11G7A,
        "D12":D12,
        "D12N":D12N,
        "D12G1C":D12G1C,
        "D12G1A":D12G1A,
        "D12G2C":D12G2C,
        "D12G2A":D12G2A,
        "D12G3C":D12G3C,
        "D12G3A":D12G3A,
        "D12G4C":D12G4C,
        "D12G4A":D12G4A,
        "D12G5C":D12G5C,
        "D12G5A":D12G5A,
        "D12G6C":D12G6C,
        "D12G6A":D12G6A,
        "D12G7C":D12G7C,
        "D12G7A":D12G7A,
        "D13":D13,
        "D13N":D13N,
        "D13G1C":D13G1C,
        "D13G1A":D13G1A,
        "D13G2C":D13G2C,
        "D13G2A":D13G2A,
        "D13G3C":D13G3C,
        "D13G3A":D13G3A,
        "D13G4C":D13G4C,
        "D13G4A":D13G4A,
        "D13G5C":D13G5C,
        "D13G5A":D13G5A,
        "D13G6C":D13G6C,
        "D13G6A":D13G6A,
        "D13G7C":D13G7C,
        "D13G7A":D13G7A,
        "D14":D14,
        "D14N":D14N,
        "D14G1C":D14G1C,
        "D14G1A":D14G1A,
        "D14G2C":D14G2C,
        "D14G2A":D14G2A,
        "D14G3C":D14G3C,
        "D14G3A":D14G3A,
        "D14G4C":D14G4C,
        "D14G4A":D14G4A,
        "D14G5C":D14G5C,
        "D14G5A":D14G5A,
        "D14G6C":D14G6C,
        "D14G6A":D14G6A,
        "D14G7C":D14G7C,
        "D14G7A":D14G7A,
        "D15":D15,
        "D15N":D15N,
        "D15G1C":D15G1C,
        "D15G1A":D15G1A,
        "D15G2C":D15G2C,
        "D15G2A":D15G2A,
        "D15G3C":D15G3C,
        "D15G3A":D15G3A,
        "D15G4C":D15G4C,
        "D15G4A":D15G4A,
        "D15G5C":D15G5C,
        "D15G5A":D15G5A,
        "D15G6C":D15G6C,
        "D15G6A":D15G6A,
        "D15G7C":D15G7C,
        "D15G7A":D15G7A,
        "D16":D10,
        "D16N":D16N,
        "D16G1C":D16G1C,
        "D16G1A":D16G1A,
        "D16G2C":D16G2C,
        "D16G2A":D16G2A,
        "D16G3C":D16G3C,
        "D16G3A":D16G3A,
        "D16G4C":D16G4C,
        "D16G4A":D16G4A,
        "D16G5C":D16G5C,
        "D16G5A":D16G5A,
        "D16G6C":D16G6C,
        "D16G6A":D16G6A,
        "D16G7C":D16G7C,
        "D16G7A":D16G7A,
        "E1N":E1N,
        "E1C":E1C,
        "E1A":E1A,
        "E2N":E2N,
        "E2C":E2C,
        "E2A":E2A,
        "E3N":E3N,
        "E3C":E3C,
        "E3A":E3A,
        "E4N":E4N,
        "E4C":E4C,
        "E4A":E4A,
        "E5N":E5N,
        "E5C":E5C,
        "E5A":E5A,
        "E6N":E6N,
        "E6C":E6C,
        "E6A":E6A,
        "E7N":E7N,
        "E7C":E7C,
        "E7A":E7A,
        "E8N":E8N,
        "E8C":E8C,
        "E8A":E8A,
        "E9N":E9N,
        "E9C":E9C,
        "E9A":E9A,
        "E10N":E10N,
        "E10C":E10C,
        "E10A":E10A,
        "E11N":E11N,
        "E11C":E11C,
        "E11A":E11A,
        "E12N":E12N,
        "E12C":E12C,
        "E12A":E12A,
        "E1":Elist[0],
        "E2":Elist[1],
        "E3":Elist[2],
        "E4":Elist[3],
        "E5":Elist[4],
        "E6":Elist[5],
        "E7":Elist[6],
        "E8":Elist[7],
        "E9":Elist[8],
        "E10":Elist[9],
        "E11":Elist[10],
        "E12":Elist[11],
        "test":CE1

        }
    return render(request, "blog/ADC22ESTA.html",context=params)


def upload(request):
    data='nothing'
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    my_file = os.path.join(THIS_FOLDER,'psw.txt')
    f = open(my_file,'r', encoding='UTF-8')
    data = f.read()
    data = data.replace('\n',',')
    data = data.split(',')
    BB = len(data)
    CC = int(BB/5)
    if request.method == 'POST':
        for i in range(CC) :
            j=5*i
            k=j+1
            l=j+2
            m=j+3
            n=j+4
            AA = User()
            AA.username = data[j]
            AA.password = data[k]
            AA.first_name = data[l]
            AA.last_name = data[m]
            AA.email = data[n]
            AA.save()
            AA.set_password(str(data[k]))
            AA.save()
        object_list = User.objects.all()
        for o in object_list:
            add_user = User.objects.filter(username=o.username).first()
            if o.username != 'hyumanozaki':
                if o.username != 'hyuma.nozaki@cosmotec.com.br':
                    BB = MBO22()
                    BB.user = add_user
                    BB.save()
                    CC = ADC22()
                    CC.user = add_user
                    CC.save()
                    DD = PDI22()
                    DD.user = add_user
                    DD.save()                   

        return redirect(to='/RH')
    params = {"form":data,"BB":BB}
    return render(request, 'blog/upload.html', params)

def upMBO(request):
    data='nothing'
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    my_file = os.path.join(THIS_FOLDER,'upM.txt')
    f = open(my_file,'r', encoding='UTF-8')
    data = f.read()
    data = data.replace('\n','^')
    data = data.split('^')
    BB = len(data)
    CC = int(BB)
    object_list = User.objects.all()
    if request.method == 'POST':
        for i in range(CC) :
            k=1
            for j in object_list:
                if data[i] == j.username:
                    task = get_object_or_404(MBO22, user=j.id)
                    task.MBO22A1=data[i+1]
                    if data[i+2] != '#':
                        task.MBO22B1=data[i+2]
                        k+=1
                        if data[i+3] != '#':
                            task.MBO22C1=data[i+3]
                            k+=1
                            if data[i+4] != '#':
                                task.MBO22D1=data[i+4]
                                k+=1
                                if data[i+5] != '#':
                                    task.MBO22E1=data[i+5]
                                    k+=1
                                    if data[i+6] != '#':
                                        task.MBO22F1=data[i+6]
                                        k+=1
                                        if data[i+7] != '#':
                                            task.MBO22G1=data[i+7]
                                            k+=1
                    if k==1:
                        task.MBO22AP=data[i+2+k]
                    if k==2:
                        task.MBO22AP=data[i+2+k]
                        task.MBO22BP=data[i+2+k+1]
                    if k==3:
                        task.MBO22AP=data[i+2+k]
                        task.MBO22BP=data[i+2+k+1]
                        task.MBO22CP=data[i+2+k+2]
                    if k==4:
                        task.MBO22AP=data[i+2+k]
                        task.MBO22BP=data[i+2+k+1]
                        task.MBO22CP=data[i+2+k+2]
                        task.MBO22DP=data[i+2+k+3]
                    if k==5:
                        task.MBO22AP=data[i+2+k]
                        task.MBO22BP=data[i+2+k+1]
                        task.MBO22CP=data[i+2+k+2]
                        task.MBO22DP=data[i+2+k+3]
                        task.MBO22EP=data[i+2+k+4]
                    if k==6:
                        task.MBO22AP=data[i+2+k]
                        task.MBO22BP=data[i+2+k+1]
                        task.MBO22CP=data[i+2+k+2]
                        task.MBO22DP=data[i+2+k+3]
                        task.MBO22EP=data[i+2+k+4]
                        task.MBO22FP=data[i+2+k+5]
                    if k==7:
                        task.MBO22AP=data[i+2+k]
                        task.MBO22BP=data[i+2+k+1]
                        task.MBO22CP=data[i+2+k+2]
                        task.MBO22DP=data[i+2+k+3]
                        task.MBO22EP=data[i+2+k+4]
                        task.MBO22FP=data[i+2+k+5]
                        task.MBO22GP=data[i+2+k+6]
                    task.save()
                    break                  
        return redirect(to='/RH')
    params = {"form":data,"BB":BB}
    return render(request, 'blog/upMBO.html', params)



def upADC(request):

    today = datetime.now()
    Y = today.year
    M = today.month
    D = today.day

    data='nothing'
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    my_file = os.path.join(THIS_FOLDER,'upA.txt')
    f = open(my_file,'r', encoding='UTF-8')
    data = f.read()
    data = data.replace('\n','^')
    data = data.split('^')
    BB = len(data)
    CC = int(BB/28)
    object_list = User.objects.all()    
    if request.method == 'POST':
        for i in range(CC) :
            k=1
            l=i*28
            for j in object_list:
                if data[l] == j.username:
                    task = get_object_or_404(ADC22, user=j.id)
                    for m in range(18):
                        if data[l+m+1]=='None':
                            data[l+1+m]=1.0
                    task.ADC22G1C=float(data[l+1])
                    task.ADC22G2C=float(data[l+2])
                    task.ADC22G3C=float(data[l+3])
                    task.ADC22G4C=float(data[l+4])
                    task.ADC22G5C=float(data[l+5])
                    task.ADC22G6C=float(data[l+6])
                    task.ADC22G7C=float(data[l+7])
                    task.ADC22E1C=float(data[l+8])
                    task.ADC22E2C=float(data[l+9])
                    task.ADC22G1A=float(data[l+10])
                    task.ADC22G2A=float(data[l+11])
                    task.ADC22G3A=float(data[l+12])
                    task.ADC22G4A=float(data[l+13])
                    task.ADC22G5A=float(data[l+14])
                    task.ADC22G6A=float(data[l+15])
                    task.ADC22G7A=float(data[l+16])
                    task.ADC22E1A=float(data[l+17])
                    task.ADC22E2A=float(data[l+18])
                    task.ADC22G1O=data[l+19]
                    task.ADC22G2O=data[l+20]
                    task.ADC22G3O=data[l+21]
                    task.ADC22G4O=data[l+22]
                    task.ADC22G5O=data[l+23]
                    task.ADC22G6O=data[l+24]
                    task.ADC22G7O=data[l+25]
                    task.ADC22E1O=data[l+26]
                    task.ADC22E2O=data[l+27]
                    task.ADC22E1=j.first_name
                    task.ADC22E2=j.last_name
                    task.ADC22C=1
                    task.ADC22A=1
                    task.ADC22Y=Y
                    task.ADC22M=M
                    task.ADC22D=D
                    task.save()
                    break                  
        return redirect(to='/RH')
    XX=BB/28
    params = {"form":data,"BB":XX}
    return render(request, 'blog/upADC.html', params)


def upPDI(request):

    today = datetime.now()
    Y = today.year
    M = today.month
    D = today.day
    
    data='nothing'
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    my_file = os.path.join(THIS_FOLDER,'upP.txt')
    f = open(my_file,'r', encoding='UTF-8')
    data = f.read()
    data = data.replace('\n','^')
    data = data.split('^')
    BB = len(data)
    CC = int(BB/14)
    object_list = User.objects.all()
    if request.method == 'POST':
        for i in range(CC) :
            k=1
            l=i*14
            for j in object_list:
                if data[l] == j.username:
                    task = get_object_or_404(PDI22, user=j.id)
                    task.PDI22G1C=data[l+1]
                    task.PDI22G2C=data[l+2]
                    task.PDI22G3C=data[l+3]
                    task.PDI22G1PD=data[l+4]
                    task.PDI22G2PD=data[l+5]
                    task.PDI22G3PD=data[l+6]
                    task.PDI22E1PD=data[l+7]
                    task.PDI22E2PD=data[l+8]
                    task.PDI22G1=data[l+9]
                    task.PDI22G2=data[l+10]
                    task.PDI22G3=data[l+11]
                    task.PDI22E1=data[l+12]
                    task.PDI22E2=data[l+13]
                    task.PDI22E1C=j.first_name
                    task.PDI22E2C=j.last_name
                    task.PDI22C=1
                    task.PDI22A=1
                    task.PDI22Y=Y
                    task.PDI22M=M
                    task.PDI22D=D
                    task.save()
                    break                  
        return redirect(to='/RH')
    XX=BB/14
    params = {"form":data,"BB":XX}
    return render(request, 'blog/upPDI.html', params)


def upCE(request):
    object_list = User.objects.all()
    if request.method == 'POST':
        for i in object_list:
            task = get_object_or_404(ADC22, user=i.id)
            task.ADC22E1=i.first_name
            task.ADC22E2=i.last_name
            task.save()
        return redirect(to='/RH')
    return render(request, 'blog/upCE.html')

def upcomp(request):

    if request.method == 'POST':
        task = get_object_or_404(RHDT, user=request.user)
        task.ADC22G1='Foco em Resultados'
        task.ADC22G1D='Foco em Resultados: capacidade de identificar e priorizar o que importa para gerar alto impacto nos resultados, entendendo e contribuindo para execução da estratégia da empresa, multiplicando seus conceitos e implementando ações para superar metas definidas para sua área e para empresa.'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='1º Nível de Proficiência : Não atinge o nível mínimo de proficiência'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='2º Nível de Proficiência: Dedicado - Estabelece e mantém bom nível de desempenho entendendo as bases da estratégia da empresa. Organiza as tarefas de acordo com as prioridades estabelecidas e demonstra iniciativa e persistência na sua execução. Contribui com sugestões práticas e efetivas.'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='3º Nível de Proficiência: Busca metas desafiadoras – Contribui fortemente para o atingimento das metas da sua área e demostras bom nível de conhecimento da estratégia e dos objetivos da empresa. Coloca metas pessoais desafiadoras e trabalha para alcançá-las. Assume responsabilidade adicional se necessário e antecipa problemas que possam ocorrer e busca soluções.'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='4º Nível de Proficiência: Impacta o resultado da sua diretoria - Contribui fortemente para o atingimento das metas da sua Diretoria e atinge resultados superiores. Tem profundo entendimento e atua como multiplicador da estratégia e objetivos da empresa. Encontra alternativas para superar os obstáculos. É capaz de priorizar suas atividades com foco no atingimento de resultados.'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='\n'
        task.ADC22G1D+='5º Nível de Proficiência: Impacta o negócio - Mobiliza recursos e obtém a cooperação/adesão de outros para atingir os objetivos com alto impacto no resultado da Empresa. Coloca os objetivos organizacionais acima das metas da área ou individuais. É tido como uma força propulsora pela organização e uma referência no conhecimento e até no desenvolvimento de estratégias para empresa.'
        task.ADC22G2='Foco no Cliente'
        task.ADC22G2D='Foco no Cliente: Capacidade de perceber e entender os requisitos e necessidades dos clientes internos e externos, gerando soluções de alto valor agregado alinhadas aos objetivos e estratégias da empresa.'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='1º Nível de Proficiência : Não atinge o nível mínimo de proficiência'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='2º Nível de Proficiência: Faz follow-up - Monitora a satisfação dos Clientes, mantém comunicação clara, responde às suas necessidades e fornece informações adequadas e precisas'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='3º Nível de Proficiência: Assume responsabilidade – Conhece as necessidades dos clientes e mantém constante interação. Nos momentos críticos, assume a responsabilidade pela correção dos problemas, procurando as melhores alternativas tanto para o cliente como para empresa'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='4º Nível de Proficiência: Agrega valor – Antecipa as necessidades dos clientes, gera soluções que agregam valor ao seu negócio e supera suas expectativas. É capaz de alinhar as ações e soluções para os clientes aos objetivos estratégicos da empresa'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='\n'
        task.ADC22G2D+='5º Nível de Proficiência: Constrói parceria - É reconhecido como referência no atendimento ao cliente. Atua como consultor, pois conhece o negócio do cliente e da empresa, envolve-se nos seus processos decisórios gerando novas oportunidades e possibilidades de implementação de projetos, produtos e soluções de alto valor agregado tanto para a companhia como para o cliente interno ou externo.'
        task.ADC22G3='Busca por Excelência'
        task.ADC22G3D='Busca por Excelência: capacidade de identificar oportunidades de melhoria e colocar o foco na solução e não no problema, identificando e atacando a causa raiz para promover continuamente a melhoria de produtos, processos e serviços, contribuindo para que a empresa seja reconhecida como a melhor do setor'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='1º Nível de Proficiência : Não atinge o nível mínimo de proficiência'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='2º Nível de Proficiência: Padrão - Claro e preciso – Trabalha de forma precisa e em conformidade com os requisitos e regras estabelecidas. Verifica o cumprimento de cada etapa do processo/ trabalho em função de expectativas de qualidade e prazo, fazendo revisões e pontos de checagem.'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='3º Nível de Proficiência: Melhoria - Sistemático no dia-a-dia – Sabe identificar pontos de melhoria, ineficiências e age para eliminá-las. Atua na causa raiz dos problemas, propõe soluções, revisa processos e programas que garantam a melhoria contínua. Deixa claro e detalha o que fazer e como fazer.'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='4º Nível de Proficiência: Projetos de Alto Impacto – Orienta a implementação e monitoramento das iniciativas de melhoria: vai além da causa-raiz, antecipa problemas, sistematiza rotinas, libera o trabalho de outros, eliminando interferências e/ou procedimentos desnecessários. Garante que as pessoas utilizem os métodos e processos chave para execução das tarefas e procedimentos de modo eficiente.'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='5º Nível de Proficiência: Catalisador – implementa e suporta as iniciativas estratégicas para melhoria de processos com impacto relevante no negócio e de forma integrada. Soluciona problemas crônicos de alta complexidade (relacionados a processos, produtos e serviços).'
        task.ADC22G3D+='\n'
        task.ADC22G3D+='\n'
        task.ADC22G4='Liderança'
        task.ADC22G4D='Liderança: Capacidade de gerar resultados através das pessoas, projetos, ou processos, mobilizando pessoas de diferentes áreas, conseguindo envolvê-las adequadamente e influenciando-as positivamente para o atingimento dos objetivos e resultados esperados.'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='1º Nível de Proficiência : Não atinge o nível mínimo de proficiência'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='2º Nível de Proficiência: Transmite as diretrizes de atividades, processos ou projetos sob sua responsabilidade – atuando como bom exemplo na execução de suas atividades e de sua área de atuação. Sabe ouvir e busca envolver as pessoas para atingir os resultados esperados.'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='3º Nível de Proficiência: Assegura a eficácia da atuação das pessoas em ações, projetos ou processos – consegue envolver adequadamente as pessoas para atingimento dos resultados esperados de maneira eficiente e eficaz. Sabe ouvir e comunicar as razões para os caminhos e planos propostos, visando o cumprimento das diretrizes e prazos dos planos negociados.'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='4º Nível de Proficiência: Constrói relação de credibilidade junto as pessoas – influenciando-as de modo positivo, provendo apoio e reconhecimento para entregarem resultados acima do esperado nas atividades, processos ou projetos sob sua responsabilidade. Fornece e mobiliza recursos, é reconhecido como isento no julgamento e ético nas relações de trabalho. Está aberto a sugestões e críticas, sabendo ouvir e delega atividades quando necessário. Estabelece indicadores de desempenho'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='\n'
        task.ADC22G4D+='5º Nível de Proficiência: Reconhecido como líder - Motiva e energiza as pessoas, obtendo adesão para atingir resultados superiores. É formador de opinião e gera ambiente confiável de trabalho estabelecendo relações e alianças adequadas garantindo alto nível de desempenho, comprometimento e responsabilidade.'
        task.ADC22G5='Senso de Urgência'
        task.ADC22G5D='Senso de Urgência : Capacidade de agir com senso de urgência, agilidade e proatividade para responder adequadamente as situações que necessitem de ações rápidas e efetivas para soluções dos problemas. Saber lidar com mais de uma atividade e priorizá-las de maneira adequada.'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='1º Nível de Proficiência : Não atinge o nível mínimo de proficiência'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='2º Nível de Proficiência: - Protocolar - Atua conforme o protocolo da área e das atividades sob sua responsabilidade minimamente respeitando os prazos estabelecidos. Limita-se a realização das atividades do dia a dia e precisa de orientação próxima e constante para atuar em demandas novas ou que exigiam respostas rápidas.'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='3º Nível de Proficiência: Adaptável – Atua com boa eficiência e responsividade nas atividades do seu dia a dia atendendo demandas de forma ágil e efetiva. Reconhece as novas demandas e que aquelas que precisam de maior agilidade de resposta para solução de problemas, modificando seu comportamento ou procedimentos para adequar-se à nova situação. Demonstra que consegue priorizar atividades e melhora essa gestão com pequeno nível de orientação superior.'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='4º Nível de Proficiência: - Ágil - Quebra os padrões de comportamento ou procedimentos usuais para aumentar sua responsividade de maneira ágil, rápida e assertiva. Tem alto nível de prontidão e é capaz de gerenciar novas demandas juntamente com suas atividades do dia a dia com excelente capacidade de priorização e entrega.'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='\n'
        task.ADC22G5D+='5º Nível de Proficiência: Ágil e Eficaz – É visto como referência em senso de urgência, absorvendo demandas novas e complexas, respondendo-as de modo extremamente ágil e provendo soluções completas e efetivas. Propõe novos padrões de comportamentos e serve de orientador e estimulando o desenvolvimento do senso de urgência junto às pessoas.'
        task.ADC22G6='Comunicação'
        task.ADC22G6D='Comunicação: capacidade de ter uma comunicação clara e objetiva, tanto de maneira oral como escrita, associada a uma boa capacidade de ouvir ativamente seu interlocutor. Capacidade realizar uma apresentação ou expor um tema levando as ideias de forma clara e conduzindo os ouvintes para o entendimento das informações passadas e feedbacks adequados.'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='1º Nível de Proficiência : Não atinge o nível mínimo de proficiência'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='2º Nível de Proficiência: - Comunica-se igualmente com todas as pessoas, sem se preocupar com adaptação ao interlocutor e com menor eficiência na objetividade. Ainda precisa desenvolver a escuta ativa para aumentar clareza na comunicação e não perder o foco constantemente, prejudicando o conteúdo da mensagem a ser passada.'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='3º Nível de Proficiência: - Reconhece o interlocutor e procura ouvir ativamente. Mante-se atento para adaptar sua comunicação de modo a melhorar a clareza e minimizar a perda de foco e objetividade;'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='4º Nível de Proficiência: ouve atentamente o interlocutor adapta sua comunicação adequadamente, comunicando-se de forma bem estruturada e objetiva conseguindo manter o foco, mesmo quando confrontado com perguntas, de modo a garantir o bom entendimento das informações a serem passadas e abre espaço para receber feedbacks.'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='\n'
        task.ADC22G6D+='5º Nível de Proficiência: Comunicação de forma muito clara e precisa, adapta-se com facilidade ao estilo do interlocutor, sempre mantendo o foco no assunto com alto nível de objetividade associado ao um bom engajamento do interlocutor. Estabelecendo rapidamente confiança e entrega completa das informações a serem passadas e dando espaço e extraindo feedbacks de alta qualidade.'
        task.ADC22G7='Relacionamento'
        task.ADC22G7D='Relacionamento: É a capacidade de perceber os comportamentos e expectativas dos interlocutores, fomentando o trabalho em equipe por meio de posturas e atitudes construtivas na relação inter-pessoal e inter-funcional. Desenvolver bom “trânsito” e relacionamento com as áreas de interface e com pessoas de diferentes níveis hierárquicos.'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='1º Nível de Proficiência : Não atinge o nível mínimo de proficiência'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='2º Nível de Proficiência: Percebe o interlocutor e atua - Percebe as características de alguns interlocutores de contato diário, decodifica o ambiente de trabalho e adota postura transparentes, agindo com espírito de equipe.'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='3º Nível de Proficiência: Percebe o grupo - Percebe a dinâmica do grupo com que se relaciona, procura resolver os problemas que limitam a comunicação, mesmo se expressos de forma não verbal, auxiliando na resolução de conflitos ou mal-entendidos. Trabalha bem em equipe do seu grupo ou multidisciplinares.'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='4º Nível de Proficiência: Transita inter-areas - Transita facilmente em vários níveis da organização. Estabelece uma relação com facilidade com qualquer tipo de interlocutor. Institui uma relação de parceria eficaz no grupo e entre áreas/interfaces. Sabe gerenciar conflitos, fomenta e facilita o trabalho em equipe do seu grupo, multidisciplinares, mesmo com membros de diferentes níveis hierárquicos.'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='\n'
        task.ADC22G7D+='5º Nível de Proficiência: Facilitador - Envolve-se em contextos complexos e de alto conflito nos diversos estilos da relação inter-pessoal. Sabe encontrar o estilo de relacionamento mais apropriado, harmonizando as relações e facilitando os fluxos e processos na organização. É reconhecido como facilitador.'
        task.ADC22E1C='Busca de Conhecimento'
        task.ADC22E1D='Busca de Conhecimento - Conceito: Capacidade de adquirir e aplicar conhecimentos para impactar positivamente no resultado do negócio, bem como difundir o conhecimento adquirido para outros profissionais'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='2º Nível de Proficiência: Mantém-se atualizado – busca de forma independente, em sua área de atuação, cursos, leituras, estudos e participação em eventos com profissionais da área.'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='3º Nível de Proficiência: Aprofunda - quando confrontado com desafios técnicos, “vai a fundo”, não se contenta com as primeiras explicações ou evidências. Recorre a outras fontes para checar se os dados obtidos estão corretos e utiliza ferramentas de análise'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='4º Nível de Proficiência: Difunde seus conhecimentos - fornece apoio técnico a outras áreas da empresa e ministrar treinamentos. Analisa e vai à raiz dos problemas de forma sistêmica medindo os riscos e prevendo as conseqüências de determinadas conclusões (em outras áreas, processos, tecnologias).'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='\n'
        task.ADC22E1D+='5º Nível de Proficiência: Chega a conclusões "fora do comum" - É percebido como referência em sua área de atuação. Busca novas fontes e disponibiliza informações que permitam à empresa identificar novas oportunidades. É capaz de enxergar relações entre dados que os outros não conseguem.'
        task.ADC22E2C='Competência Técnica de Análise'
        task.ADC22E2D='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='2º Nível de Proficiência: Possui conhecimento básico das causas e efeitos de problemas, aplicando-os para obter soluções de problemas simples. É capaz de fazer levantamento de dados e análises de indicadores dentro de um formato pré-determinado, assegurando sua exatidão. Necessita de supervisão e/ou acompanhamento constante para analisar as informações, definir as causas dos problemas e propor soluções.'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='3º Nível de Proficiência: Demonstra conhecimentos detalhados dos relatórios e controles da sua área e de análises numéricas e estimativas. É capaz de fazer análises simplificadas para solucionar problemas, possuindo habilidade em assimilar e processar informações, identificando situações críticas, desvios ou inconsistências. Propõe soluções técnicas de modo normalizar eventuais desvios em relação as especificações e padrões.'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='4º Nível de Proficiência: É capaz de solucionar problemas, analisar a viabilidade técnica e assimilar / processar análises complexas. Possui visão sistêmica do processo de sua área. É capaz de agir de maneira preventiva, sintetizar soluções rapidamente, estabelecer recomendações e planos de ação alinhados às estratégias da empresa. Tem habilidade em propor melhorias que envolvam o desenvolvimento de novos processos e/ou implementação de técnicas / metodologias disponíveis na empresa ou no mercado.'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='\n'
        task.ADC22E2D+='5º Nível de Proficiência: Possui visão estratégica da área. É capaz de solucionar problemas complexos de maneira ágil e eficaz. Utiliza ferramentas de análise estratégica profunda para solução de problemas crônicos. Tem habilidade para definir medidas preventivas e analisar a viabilidade técnica de implementação ou revisão de processos, considerando a relação custo x benefício x oportunidade.'
        task.ADC22E3C='Competência Técnica em Negociação'
        task.ADC22E3D='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='2º Nível de Proficiência: Facilita o diálogo em situações do dia-a-dia. Envolve-se em acordos de baixa complexidade e impacto no ambiente interno de trabalho, negociando dentro de normas e procedimentos pré-existentes e sendo orientado para a busca de resultados que equilibrem expectativas das partes interessadas. Negocia internamente recursos e condições operacionais para a execução de suas atividades, seguindo diretrizes estabelecidas. Age sob supervisão direta.'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='3º Nível de Proficiência: Demonstra capacidade em propor acordos de média complexidade. Atua sob supervisão ocasional. Identifica os problemas, potenciais objeções, necessidades e proposições do outro interlocutor, utilizando informações para planejar, de forma coerente, seus argumentos e condições de negociação. É aberto para rever posições e entender pontos de vista distintos dos seus. Pode ver benefícios subjacentes que o outro interlocutor não vê e os comunica. Nas negociações, apresenta argumentação consistente e fundamentada, transmitindo credibilidade e confiança.'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='4º Nível de Proficiência: Apresenta conhecimentos sobre as variáveis de impacto e risco em suas interfaces e nos negócios. É capaz de negociar com múltiplos interlocutores externos, propondo soluções em situações de conflito e influenciando pessoas para o estabelecimento de uma relação ganha-ganha, com foco em gerar credibilidade e obter resultados de curto prazo. É capaz de estruturar negociações que envolvam mais de um interlocutor, identificando seus estilos de negociação, introduzindo novos argumentos e integrando posições divergentes. Lidera negociações externas de grande porte, de natureza não estruturada.'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='\n'
        task.ADC22E3D+='5º Nível de Proficiência: Possui habilidade de argumentação para obter a adesão e cooperação das pessoas para atingir os objetivos esperados. Constrói uma rede de relacionamento, dentro e fora da empresa, de modo a facilitar o processo de negociação, estabelecendo relações duradouras. Reconhece o modelo de negociação do outro interlocutor e é capaz de adaptar o seu visando estabelecer parcerias e alianças estratégicas. Diante de situações inesperadas, capta novos elementos, estruturando e propondo soluções e/ou ações oportunas assertivamente. Identifica os pontos fortes e fracos dos outros interlocutores e utiliza este conhecimento em sua estratégia de negociação. Tem uma ampla gama de estratégias e as aplica de acordo com a situação. É responsável por desenvolver e definir negociações que causam impacto em toda organização.'
        task.ADC22E4C='Competência Técnica em Conhecimento do Negócio'
        task.ADC22E4D='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='2º Nível de Proficiência: Possui conhecimentos básicos do negócio e noções dos produtos da Empresa, bem como uma visão geral dos padrões e procedimentos da empresa aplicados à sua área de atuação.'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='3º Nível de Proficiência: Demonstra conhecimentos das políticas, dos padrões, regras de negócios e dos produtos Empresa e suas aplicações, sendo capaz de utilizá-los para identificar oportunidades e sugerir melhorias em sua área de atuação. Tem conhecimento sobre as práticas do mercado, fornecedores e clientes.'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='4º Nível de Proficiência: Possui amplos conhecimentos das características dos produtos, processos ou mercados da Empresa, variáveis e tendências do negócio. É capaz de desenvolver idéias sobre como os produtos, processos ou tecnologias devem ser aplicados e/ou adaptados às diversas situações de mercado, demandadas em sua área. Demonstra habilidade em identificar o impacto de situações de mercado nas atividades de sua área e interfaces e em propor alternativas e ações preventivas.'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='\n'
        task.ADC22E4D+='5º Nível de Proficiência: Possui profundos conhecimentos para inferir sobre tendências de mercado futuras dentro do seu campo de atuação. Demonstra habilidade em enxergar as perspectivas do negócio e seus impactos na estratégia da empresa, objetivando manter a organização sempre informada sobre mudanças, ameaças e oportunidades. Seus conhecimentos permitem a elaboração de estratégias e a definição de tendências e regras de atuação no mercado.'
        task.ADC22E5C='Desenvolvimento de Pessoas'
        task.ADC22E5D='Desenvolvimento de Pessoas: É a capacidade de captar, treinar, desenvolver pessoas. Isto se revela pela sua postura no dia a dia com sua equipe e no seu histórico de formação de pessoas para a empresa'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='2º Nível de Proficiência: Faz acompanhamentos ocasionais - Dá orientações, mostra como se faz, fornece instruções para a execução das tarefas. Fornece feedback apenas genérico e não direcionado.'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='3º Nível de Proficiência: Dá feedback – Identifica e comunica pontos fortes e fracos das pessoas e traça plano de treinamento. Sabe ouvir e está aberto a receber feedback. Busca captar potenciais e recursos adicionais para formá-los (treinamentos, estágios, leituras).'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='4º Nível de Proficiência: Apoiador – Faz acompanhamentos constantes e orienta as pessoas para atingirem suas metas mostrando-lhes alternativas. Realiza a avaliação de desempenho dando feedback construtivo, específico e pontual. Acompanha o cumprimento do plano de treinamento e encoraja as pessoas a buscarem seu auto-aperfeiçoamento. Tem preparado bons profissionais para a área e para a empresa.'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='\n'
        task.ADC22E5D+='5º Nível de Proficiência: É desenvolvedor -Forma sucessores, faz coaching, acompanha a evolução das pessoas, oferece conselhos e apoia o crescimento profissional. Identifica oportunidades de rotatividade entre áreas. É reconhecido pela equipe e pela empresa como desenvolvedor de talentos.'
        task.ADC22E6C='Visão Estratégica'
        task.ADC22E6D='Visão Estratégica: Capacidade de entender e contribuir para estratégia da empresa, multiplicando os conceitos e implementando ações táticas, relacionadas à sua área de atuação para atingir os objetivos da empresa'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='2º Nível de Proficiência: Informado - Procura entender as metas de sua área como parte importante da estratégia da organização, gerando ações do dia a dia que contribuem para melhoria de resultados.'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='3º Nível de Proficiência: Visão tática do negócio - Entende os objetivos das iniciativas estratégicas de sua área e diretoria e propõe planos práticos para implementá-los. Busca informações externas sobre tendências do mercado, dentro de sua área de atuação, para construir planos táticos.'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='4º Nível de Proficiência: Visão Estratégica do Negócio - Elabora e implementa planos para concretizar uma visão de futuro para sua área que tenha impacto nas iniciativas estratégicas da Empresa. Define foco para ação com base na visão do todo.'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='\n'
        task.ADC22E6D+='5º Nível de Proficiência: Visão Estratégica Ampla - Tem uma visão da cadeia de valor no longo prazo e explicita possíveis impactos na estratégia da empresa. Entende as demandas da organização e as traduz em ações, realizando projetos e mobilizando recursos humanos, técnicos e físicos para sua concretização.'
        task.ADC22E7C='Inovação'
        task.ADC22E7D='Inovação: capacidade de criar, desenvolver e implementar soluções criativas e diferenciadas que agreguem valor ao resultado esperado, seja para a empresa, cliente, produto, processo ou serviço.'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='2º Nível de Proficiência: Interesse pela inovação – Contribui com idéias criativas relativas à sua atividade e de baixo impacto no negócio.'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='3º Nível de Proficiência: Gera soluções – Sugere alternativas criativas, fora dos padrões usuais, solucionando problemas antigos dentro da sua área de atuação.'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='4º Nível de Proficiência: Soluções Originais - É capaz de elaborar soluções criativas a partir de observações fora do seu meio de atuação. Em geral são soluções reconhecidas por outros como originais. Apóia e incentiva idéias criativas, testando novas abordagens que podem criar valor para a Empresa. Encoraja os outros a assumir os riscos apropriados para obter os ganhos e os benefícios de novas idéias.'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='\n'
        task.ADC22E7D+='5º Nível de Proficiência: Novos Conceitos – Consegue encontrar soluções e implementar novas idéias, de alto impacto para a Empresa, relacionadas a problemas complexos ou situações novas que exigem a criação e o desenvolvimento de conceitos originais. É capaz de raciocinar fora dos padrões convencionais. Sugere produtos ou soluções diferenciados que abrem vantagens competitivas para a Empresa.'
        task.ADC22E8C= 'Flexibilidade'
        task.ADC22E8D= 'Flexibilidade - Conceito: É a capacidade de adaptar ou mudar o seu comportamento ou estilo de trabalho diante de perspectivas diferentes e até mesmo opostas'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='2º Nível de Proficiência: Pouco flexível - Reconhece os pontos de vista diferentes do seu, mas mantém seu posicionamento e opinião.'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='3º Nível de Proficiência: Adaptável – Modifica seu comportamento ou procedimentos usuais para adequar-se a uma situação específica com objetivo de conseguir que as suas metas sejam atingidas.'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='4º Nível de Proficiência: Tem “Jogo de Cintura” - Quebra os padrões de comportamento ou procedimentos usuais para conseguir os resultados esperados. Está aberto a novas iniciativas, disponibilizando'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='\n'
        task.ADC22E8D+='5º Nível de Proficiência: Estimula a Flexibilidade - Propõe novos padrões de comportamentos sem perder a visão do todo e questionando o "status quo”. Recebe bem as mudanças e estimula outros a rever seus pontos de vista e comportamentos. tempo e energia para viabilizá-las.'
        task.ADC22E9C='Competência Técnica Específica Marketing'
        task.ADC22E9D='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='2º Nível de Proficiência: Possui conhecimentos básicos sobre os procedimentos e rotinas da área de atuação, desenvolvendo uma visão da família de produtos e principais segmentos,clientes, concorrentes e fornecedores. É capaz de acompanhar os mercados, apoiar a elaboração de estudos específicos com visão de família de produtos, sob orientação'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='3º Nível de Proficiência: Possui bons conhecimentos sobre os procedimentos, técnicas e processos da área de atuação. É capaz de acompanhar os mercados, monitorar as cadeias de suprimentos (da Empresa, de concorrentes e de produtos sucedâneos), apoiar negociações mercadológicas através da elaboração de estudos específicos de média complexidade (pacote de produtos, bidding), sob supervisão. Tem conhecimentos básicos sobre os conceitos e terminologia envolvidos nos cálculos de EVA.'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='4º Nível de Proficiência: Possui conhecimentos avançados sobre os procedimentos, técnicas e processos da área de atuação. Possui bons conhecimentos de mercados, produtos, concorrentes e clientes. É capaz de consolidar informações do processo de definição da política mensal de preços mínimos. Tem habilidade na elaboração e acompanhamento do Planejamento Estratégico e Planos de Ação para cada família de produtos (cumprimento das metas, perspectivas e correções nos casos de desvio). Tem capacidade de influenciar negociações mercadológicas nas sub-famílias de produtos sob sua responsabilidade. Tem habilidade na elaboração de estudos específicos de maior complexidade (fórmulas de preços e contratos de longo prazo) sob supervisão. Tem capacidade para elaborar previsões de demanda, visando subsidiar o planejamento estratégico, sob orientação. Tem habilidade na elaboração de cálculos de EVA para decisões de investimento, com base em premissas estabelecidas.'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='\n'
        task.ADC22E9D+='5º Nível de Proficiência: Domina os conhecimentos sobre procedimentos, técnicas e processos da área de atuação, sendo considerado referência internamente. Conhece com profundidade as características do mercado em relação aos concorrentes, clientes e seus produtos. É capaz de propor a política mensal de preços mínimos, coordenar seu processo de implantação, supervisionar a elaboração de pacotes de produtos e subsidiar o plano orçamentário anual. Tem habilidade para estabelecer premissas e parâmetros de mercado que sustentem os cálculos de EVA. Coordena, elabora e propõe o planejamento estratégico dos negócios/famílias sob sua responsabilidade.'
        task.ADC22E10C='Soluções Criativas'
        task.ADC22E10D='Soluções Criativas - Conceito : É a capacidade de desenvolver novas idéias e gerar soluções criativas que resultem em aumento da competitividade, aumento de eficiência ou geração de valor'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='1. Não atinge o nível mínimo de proficiência'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='2º Nível de Proficiência: Interesse pela inovação – Contribui com idéias criativas relativas à sua atividade e de baixo impacto no negócio.'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='3º Nível de Proficiência: Gera soluções – Sugere alternativas criativas, fora dos padrões usuais, solucionando problemas antigos dentro da sua área de atuação.'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='4º Nível de Proficiência: Soluções Originais - É capaz de elaborar soluções criativas a partir de observações fora do seu meio de atuação. Em geral são soluções reconhecidas por outros como originais. Apóia e incentiva idéias criativas, testando novas abordagens que podem criar valor para a Empresa. Encoraja os outros a assumir os riscos apropriados para obter os ganhos e os benefícios de novas idéias.'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='\n'
        task.ADC22E10D+='5º Nível de Proficiência: Novos Conceitos – Consegue encontrar soluções e implementar novas idéias, de alto impacto para a Empresa, relacionadas a problemas complexos ou situações novas que exigem a criação e o desenvolvimento de conceitos originais. É capaz de raciocinar fora dos padrões convencionais. Sugere produtos ou soluções diferenciados que abrem vantagens competitivas para a Empresa.'
        task.save()
        return redirect(to='/RH')
    return render(request, 'blog/upcomp.html')

def listAE(request):
    object_list = User.objects.all()
    data=''
    for i in object_list:
        data+=i.username
        data+='^'
        data+=i.password
        data+='^'
        data+=i.first_name
        data+='^'
        data+=i.last_name
        data+='^'
        data+=i.email
        data+='^'
    data=data.split('^')
    params = {"data":data,}
    return render(request, 'blog/listAE.html',params)
    

def listMBO(request):
    object_list = User.objects.all()
    data=''
    for i in object_list:
        data+=i.username
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22A1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22B1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22C1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22D1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22E1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22F1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22G1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22AP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22BP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22CP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22DP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22EP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22FP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22GP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22A2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22B2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22C2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22D2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22E2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22F2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22G2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22A3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22B3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22C3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22D3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22E3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22F3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22G3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22A4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22B4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22C4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22D4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22E4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22F4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22G4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22AR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22BR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22CR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22DR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22ER').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22FR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO22GR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23A1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23B1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23C1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23D1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23E1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23F1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23G1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23AP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23BP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23CP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23DP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23EP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23FP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23GP').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23A2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23B2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23C2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23D2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23E2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23F2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23G2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23A3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23B3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23C3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23D3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23E3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23F3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23G3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23A4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23B4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23C4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23D4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23E4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23F4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23G4').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23AR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23BR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23CR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23DR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23ER').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23FR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=MBO22.objects.values_list('MBO23GR').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
    data = data.replace('\n','')
    data = data.split('^')
    params={"data":data}
    return render(request, 'blog/listMBO.html',params)

def listADC(request):
    object_list = User.objects.all()
    data=''
    for i in object_list:
        data+=i.username
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G3C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G4C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G5C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G6C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G7C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G1A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G2A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G3A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G4A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G5A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G6A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G7A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G1OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G2OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G3OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G4OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G5OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G6OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G7OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA+=ADC22.objects.values_list('ADC22G1O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G2O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G3O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G4O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G5O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G6O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22G7O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E1A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E2A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E1OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E2OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E1O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC22E2O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G3C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G4C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G5C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G6C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G7C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G1A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G2A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G3A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G4A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G5A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G6A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G7A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G1OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G2OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G3OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G4OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G5OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G6OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G7OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA+=ADC22.objects.values_list('ADC23G1O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G2O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G3O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G4O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G5O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G6O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23G7O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E1A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E2A').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E1OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E2OC').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E1O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=ADC22.objects.values_list('ADC23E2O').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'       
    data = data.replace('\n','')
    data = data.split('^')
    params={"data":data}
    return render(request, 'blog/listMBO.html',params)


def listPDI(request):
    object_list = User.objects.all()
    data=''
    for i in object_list:
        data+=i.username
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G1PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G2PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G3C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G3PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22G3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22E1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22E1PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22E1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22E2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22E2PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI22E2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G1PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G2PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G3C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G3PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23G3').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23E1C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23E1PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23E1').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23E2C').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23E2PD').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
        AA=[]
        AA+=PDI22.objects.values_list('PDI23E2').get(user=i.id)
        if str(AA[0])=='':
            AA[0]='None'
        data+=str(AA[0])
        data+='^'
    data = data.replace('\n','')
    data = data.split('^')
    params={"data":data}
    return render(request, 'blog/listPDI.html',params)

